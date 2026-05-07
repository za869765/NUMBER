# -*- coding: utf-8 -*-
"""
HIV 匿名諮詢代碼批次取號工具
hiva.cdc.gov.tw 自動化
衛生所外展用 — 協助民眾批次產生諮詢代碼

v1.0.0
"""
import os
import sys
import time
import random
import threading
import queue
import json
import hashlib
import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

VERSION = "1.0.61"
DEBUG = False  # v1.0.38：正式版預設關閉，失敗時 HTML 快照不再自動存

# v1.0.39 雲端授權服務（Cloudflare Worker URL）
# 部署 cloud_auth/ 後得到的 workers.dev 網址貼這
CLOUD_AUTH_URL = "https://hiv-auth.za869765.workers.dev"
CLOUD_AUTH_TIMEOUT = 8  # 秒

# ── v1.0.55 自動更新 ──
# 啟動時背景查 /version，如果版本比自己新 → 下載新 EXE 到同層、log 通知使用者
# manifest.json + latest.exe 存在 R2 bucket "hiv-exe"
def _semver_compare(a, b):
    """比較版本字串：x.y.z；回 1=a>b、-1=a<b、0=eq"""
    def parts(s):
        out = []
        for x in str(s).split("."):
            num = ""
            for c in x:
                if c.isdigit(): num += c
                else: break
            out.append(int(num) if num else 0)
        while len(out) < 3:
            out.append(0)
        return out[:3]
    pa, pb = parts(a), parts(b)
    for x, y in zip(pa, pb):
        if x > y: return 1
        if x < y: return -1
    return 0

def check_and_download_update(log_callback=None):
    """v1.0.55：背景查 /version，有新版自動下載到 EXE 同層
    log_callback: 收到狀態訊息的 callback（如 self.log；會 thread-safe 透過 root.after 呼叫）
    開發 .py 模式跳過（不下載）。"""
    def _log(m):
        if log_callback:
            try: log_callback(m)
            except Exception: pass

    if not getattr(sys, "frozen", False):
        return  # .py 開發模式跳過

    import urllib.request, urllib.error
    try:
        req = urllib.request.Request(
            CLOUD_AUTH_URL.rstrip("/") + "/version",
            headers={"User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)"},
        )
        with urllib.request.urlopen(req, timeout=CLOUD_AUTH_TIMEOUT) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if not data.get("ok"):
            return  # 還沒上傳任何版本
        latest = (data.get("latest") or "").strip()
        if not latest:
            return
        cmp_ = _semver_compare(latest, VERSION)
        if cmp_ <= 0:
            return  # 已是最新或更舊（dev 環境）
        # 找下載 URL（相對路徑）
        download_url = data.get("download_url") or "/exe-download"
        if not download_url.startswith("http"):
            download_url = CLOUD_AUTH_URL.rstrip("/") + download_url
        filename = data.get("filename") or f"HIV取號_v{latest}.exe"

        exe_dir = os.path.dirname(sys.executable)
        new_exe_path = os.path.join(exe_dir, filename)
        if os.path.exists(new_exe_path):
            _log(f"🔔 新版 v{latest} 已下載過：{filename}（請關閉本工具切換至新版）")
            return
        _log(f"🔔 偵測到新版 v{latest}（目前 v{VERSION}），背景下載中…")
        tmp_path = new_exe_path + ".downloading"
        try:
            req2 = urllib.request.Request(
                download_url,
                headers={"User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)"},
            )
            with urllib.request.urlopen(req2, timeout=120) as resp:
                with open(tmp_path, "wb") as f:
                    while True:
                        chunk = resp.read(64 * 1024)
                        if not chunk:
                            break
                        f.write(chunk)
            os.replace(tmp_path, new_exe_path)
            _log(f"✅ 新版 v{latest} 已下載：{filename}")
            _log(f"   → 請關閉本工具，將舊 EXE 刪除後雙擊新 EXE")
        except Exception as e:
            try: os.remove(tmp_path)
            except Exception: pass
            _log(f"⚠ 下載新版失敗：{e}")
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return  # 還沒上傳，安靜
        # 其他錯誤靜默
    except Exception:
        pass  # 網路問題等，靜默

# ── v1.0.28：lazy import selenium → 啟動加速 ──
# 不在 import 階段載入 selenium（拖慢 EXE 冷啟動約 1-2 秒）
# 真正用到時才在 HivaWorker._import_selenium() 裡載入
# PyInstaller 端透過 spec 的 hidden_imports 列舉確保打包完整

# ── 常數 ──────────────────────────────────────────
URL_HOME = "https://hiva.cdc.gov.tw/Default.aspx"
URL_QUESTIONNAIRE = "https://hiva.cdc.gov.tw/Questionnaire.aspx"

CITIES = [
    "台北市", "新北市", "桃園市", "台中市", "台南市", "高雄市",
    "基隆市", "新竹市", "嘉義市",
    "新竹縣", "苗栗縣", "彰化縣", "南投縣", "雲林縣", "嘉義縣",
    "屏東縣", "宜蘭縣", "花蓮縣", "台東縣",
    "澎湖縣", "金門縣", "連江縣",
]
GENDERS = ["男", "女", "跨性別", "其他"]
NATIONS = ["本國籍", "外國籍"]
EDUS    = ["不識字", "國中以下", "高中職", "專科或大學", "研究所(含)以上"]
ORIENTS = ["同性", "雙性", "異性"]
TESTING = ["否", "從未做過"]

def _default_output_dir():
    """v1.0.36：預設輸出目錄
       - EXE：與 EXE 同層資料夾（可攜，移到任何電腦都自動指向 EXE 旁邊）
       - 開發環境（.py）：沿用 D:\\Backup\\Desktop\\CODE\\number 舊路徑"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return r"D:\Backup\Desktop\CODE\number"

OUTPUT_DIR = _default_output_dir()  # v1.0.15：可由 UI 動態變更（settings.json 持久化）
DEFAULT_OUTPUT_DIR = OUTPUT_DIR

def _settings_path():
    """settings.json 永遠放在 EXE 旁邊（地端習慣），不跟著 output_dir 移動"""
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
           else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "settings.json")

SETTINGS_FILE = _settings_path()

def set_output_dir(new_path):
    """更新全域 OUTPUT_DIR；自動建立資料夾結構（含 old / log / debug）"""
    global OUTPUT_DIR
    OUTPUT_DIR = new_path
    ensure_outdir()

# ── v1.0.18 主題配色 / v1.0.47 套用 Claude Design 概念稿（Linear+Apple Health+5%AC）──
# 4 套色票對應使用情境：
#   Ocean   上午門診    清醒可信任，海平線藍配冷灰底
#   Sunset  下午外展尾聲 黃昏赤陶，溫暖收束
#   Forest  戶外場域    苔綠杉林，平靜低疲勞
#   Dark    夜間站點    深石板灰（非全黑）+ 反光衣藍
THEMES = {
    "Ocean (海洋藍)": {
        "bg": "#f3f5f8", "panel": "#ffffff", "fg": "#1a2230",
        "accent": "#2c6bd1", "accent_hover": "#1f4f9c",
        "success": "#2f7a4d", "warn": "#b07414", "error": "#b13a2e",
        "log_bg": "#14181f", "log_fg": "#c4cbd6",
        "log_success": "#5fc38a", "log_warn": "#e9b25c", "log_error": "#e87767",
    },
    "Sunset (夕陽橘)": {
        "bg": "#faf2ec", "panel": "#ffffff", "fg": "#2a201a",
        "accent": "#d96b3a", "accent_hover": "#a84c20",
        "success": "#7a8c3f", "warn": "#c08a1a", "error": "#a84830",
        "log_bg": "#14181f", "log_fg": "#c4cbd6",
        "log_success": "#5fc38a", "log_warn": "#e9b25c", "log_error": "#e87767",
    },
    "Forest (森林綠)": {
        "bg": "#f1f6f1", "panel": "#ffffff", "fg": "#1a2a1f",
        "accent": "#1f5a37", "accent_hover": "#2f7a4d",
        "success": "#2f7a4d", "warn": "#a87a1c", "error": "#a23e2b",
        "log_bg": "#14181f", "log_fg": "#c4cbd6",
        "log_success": "#5fc38a", "log_warn": "#e9b25c", "log_error": "#e87767",
    },
    "Dark (深色)": {
        "bg": "#1a1f27", "panel": "#232a35", "fg": "#e3e8ef",
        "accent": "#4a90e2", "accent_hover": "#76aef0",
        "success": "#5fc38a", "warn": "#e9b25c", "error": "#e87767",
        "log_bg": "#14181f", "log_fg": "#c4cbd6",
        "log_success": "#5fc38a", "log_warn": "#e9b25c", "log_error": "#e87767",
    },
}
DEFAULT_THEME = "Ocean (海洋藍)"

# v1.0.21 完整模式 schema（41 欄）
# 每筆 record 在 xlsx 裡的欄位順序、key、label、預設、可選值
COMPLETE_FIELDS = [
    # P1
    ("q1_sex",         "P1Q1 過性行為",           "否",     ["是", "否"]),
    ("q2_condom",      "P1Q2 全程保險套",         "沒有發生",["是", "否", "沒有發生"]),
    ("q3_regular",     "P1Q3 跟固定性伴侶",       "沒有發生",["是", "否", "沒有發生"]),
    ("q4_alcohol",     "P1Q4 用酒",               "沒有發生",["是", "否", "沒有發生"]),
    ("q5_drug",        "P1Q5 用藥",               "沒有發生",["是", "否", "沒有發生"]),
    # P2
    ("q6_std",         "P2Q6 感染性病",           "否",     ["是", "否"]),
    ("q6_hiv",         "P2Q6.1 HIV",              "否",     ["是", "否"]),
    ("q6_hiv_reason",  "P2Q6.1 HIV 想再篩原因",   "",       None),
    ("q6_warts",       "P2Q6.2 菜花",             "否",     ["是", "否"]),
    ("q6_syphilis",    "P2Q6.3 梅毒",             "否",     ["是", "否"]),
    ("q6_gonorrhea",   "P2Q6.4 淋病",             "否",     ["是", "否"]),
    ("q6_chlamydia",   "P2Q6.5 披衣菌",           "否",     ["是", "否"]),
    ("q6_herpes",      "P2Q6.6 疱疹",             "否",     ["是", "否"]),
    ("q6_hepA",        "P2Q6.7 A肝",              "否",     ["是", "否"]),
    ("q6_hepC",        "P2Q6.8 C肝",              "否",     ["是", "否"]),
    ("q6_other",       "P2Q6.9 其他性病",         "",       None),
    # P3
    ("q7_drug_use",    "P3Q7 成癮藥物",           "否",     ["是", "否"]),
    ("q7_amph",        "P3Q7.1 安非他命",         "否",     ["是", "否"]),
    ("q7_amph_method", "P3Q7.1 安非他命使用方式", "",       None),  # 多選用逗號分隔：吸入,注射,口服
    ("q7_ghb",         "P3Q7.2 G水",              "否",     ["是", "否"]),
    ("q7_mdma",        "P3Q7.3 搖頭丸",           "否",     ["是", "否"]),
    ("q7_ketamine",    "P3Q7.4 K他命",            "否",     ["是", "否"]),
    ("q7_rush",        "P3Q7.5 RUSH",             "否",     ["是", "否"]),
    ("q7_meph",        "P3Q7.6 喵喵",             "否",     ["是", "否"]),
    ("q7_heroin",      "P3Q7.7 海洛因",           "否",     ["是", "否"]),
    ("q7_marijuana",   "P3Q7.8 大麻",             "否",     ["是", "否"]),
    ("q7_other",       "P3Q7.9 其他藥物",         "",       None),
    ("q7_status",      "P3Q7.10 目前使用狀態",    "",       ["還在使用", "已停用"]),
    # P4
    ("q8a_online",     "P4Q8a 網路認識",          "否",     ["是", "否"]),
    ("q8b_venue",      "P4Q8b 娛樂場所認識",      "否",     ["是", "否"]),
    ("q8c_sex_worker", "P4Q8c 性交易服務者",      "否",     ["是", "否"]),
    ("q8d_sex_consumer","P4Q8d 性交易消費者",     "否",     ["是", "否"]),
    # P5
    ("q9_partner_hiv", "P5Q9 固定伴侶HIV",        "否",     ["是", "否", "不確定", "目前沒有固定性伴侶"]),
    ("q10_pep_used",   "P5Q10 PEP使用過",         "否",     ["是", "否"]),
    ("q11_pep_want",   "P5Q11 PEP想服用",         "否",     ["是", "否"]),
    ("q12_prep_heard", "P5Q12 聽過PrEP",          "否",     ["是", "否"]),
    ("q13_prep_want",  "P5Q13 PrEP想服用",        "否",     ["是", "否"]),
    # P6
    ("gender",         "P6 性別",                 "男",     GENDERS),
    ("nation",         "P6 國籍",                 "本國籍",  NATIONS),
    ("year",           "P6 出生年",               1990,     None),
    ("res18",          "P6 18歲前居住地",         "台南市",  CITIES),
    ("resCur",         "P6 現居住地",             "台南市",  CITIES),
    ("orient",         "P6 性傾向",               "異性",   ORIENTS),
    ("edu",            "P6 教育程度",             "高中職",  EDUS),
    # P7
    ("testing_habit",  "P7Q1 篩檢習慣",           "否",     ["是", "否", "從未做過"]),
    ("phone",          "P7 手機號碼(選填)",       "",       None),
    ("email",          "P7 E-mail(選填)",         "",       None),
    ("other_contact",  "P7 其他聯絡(選填)",       "",       None),
]
COMPLETE_KEYS = [k for k, _, _, _ in COMPLETE_FIELDS]
COMPLETE_FIELD_ALLOWED = {k: vals for k, _, _, vals in COMPLETE_FIELDS if vals}

# v1.0.29：Excel 輸出用的乾淨欄位名（無 P1Q1 prefix），對應 COMPLETE_FIELDS 的 key
OUTPUT_LABELS = {
    "q1_sex": "Q1 過性行為", "q2_condom": "Q2 全程保險套",
    "q3_regular": "Q3 跟固定性伴侶", "q4_alcohol": "Q4 用酒", "q5_drug": "Q5 用藥",
    "q6_std": "Q6 感染性病", "q6_hiv": "Q6.1 HIV", "q6_hiv_reason": "Q6.1 HIV原因",
    "q6_warts": "Q6.2 菜花", "q6_syphilis": "Q6.3 梅毒",
    "q6_gonorrhea": "Q6.4 淋病", "q6_chlamydia": "Q6.5 披衣菌",
    "q6_herpes": "Q6.6 疱疹", "q6_hepA": "Q6.7 A肝",
    "q6_hepC": "Q6.8 C肝", "q6_other": "Q6.9 其他性病",
    "q7_drug_use": "Q7 成癮藥物", "q7_amph": "Q7.1 安非他命",
    "q7_amph_method": "Q7.1 使用方式", "q7_ghb": "Q7.2 G水",
    "q7_mdma": "Q7.3 搖頭丸", "q7_ketamine": "Q7.4 K他命",
    "q7_rush": "Q7.5 RUSH", "q7_meph": "Q7.6 喵喵",
    "q7_heroin": "Q7.7 海洛因", "q7_marijuana": "Q7.8 大麻",
    "q7_other": "Q7.9 其他藥物", "q7_status": "Q7 目前使用狀態",
    "q8a_online": "Q8a 網路認識", "q8b_venue": "Q8b 娛樂場所認識",
    "q8c_sex_worker": "Q8c 性交易服務者", "q8d_sex_consumer": "Q8d 性交易消費者",
    "q9_partner_hiv": "Q9 固定伴侶HIV", "q10_pep_used": "Q10 PEP使用過",
    "q11_pep_want": "Q11 PEP想服用", "q12_prep_heard": "Q12 聽過PrEP",
    "q13_prep_want": "Q13 PrEP想服用",
    "gender": "性別", "nation": "國籍", "year": "出生年",
    "res18": "18歲前居住地", "resCur": "現居住地",
    "orient": "性傾向", "edu": "教育程度",
    "testing_habit": "篩檢習慣", "phone": "手機", "email": "E-mail",
    "other_contact": "其他聯絡",
}
# v1.0.35：摘要欄（B方案）
OUTPUT_SUMMARY_LABELS = ["篩檢路徑", "風險摘要"]
# 輸出欄位順序：基本元資料 + 摘要欄 + COMPLETE_FIELDS 順序
OUTPUT_HEADERS = (["#", "諮詢代碼", "取得時間"] + OUTPUT_SUMMARY_LABELS
                  + [OUTPUT_LABELS.get(k, k) for k in COMPLETE_KEYS])


def _classify_record(r):
    """v1.0.35：根據答案判斷篩檢路徑 + 風險摘要"""
    is_yes_q1 = r.get("q1_sex") == "是"
    is_yes_q6 = r.get("q6_std") == "是"
    is_yes_q7 = r.get("q7_drug_use") == "是"
    # 高風險指標
    risk_high = (is_yes_q6 or is_yes_q7
                 or r.get("q7_amph") == "是" or r.get("q7_heroin") == "是"
                 or r.get("q9_partner_hiv") == "是"
                 or r.get("q8a_online") == "是" or r.get("q8b_venue") == "是"
                 or r.get("q8c_sex_worker") == "是" or r.get("q8d_sex_consumer") == "是")
    # 路徑判定
    if is_yes_q6:
        path = "感染史"
    elif is_yes_q7:
        path = "藥物史"
    elif is_yes_q1:
        path = "性行為"
    else:
        path = "簡單"
    # 風險摘要
    if risk_high:
        risk = "高"
    elif is_yes_q1:
        risk = "中"
    else:
        risk = "低"
    return path, risk

# 各頁的 radio group 對應 profile keys（依 DOM 出現順序）
PAGE1_KEYS = ["q1_sex", "q2_condom", "q3_regular", "q4_alcohol", "q5_drug"]
PAGE2_BASE_KEY  = "q6_std"
PAGE2_SUB_KEYS  = ["q6_hiv", "q6_warts", "q6_syphilis", "q6_gonorrhea",
                   "q6_chlamydia", "q6_herpes", "q6_hepA", "q6_hepC"]
PAGE3_BASE_KEY  = "q7_drug_use"
PAGE3_SUB_KEYS  = ["q7_amph", "q7_ghb", "q7_mdma", "q7_ketamine", "q7_rush",
                   "q7_meph", "q7_heroin", "q7_marijuana", "q7_status"]
PAGE4_KEYS = ["q8a_online", "q8b_venue", "q8c_sex_worker", "q8d_sex_consumer"]
PAGE5_KEYS = ["q9_partner_hiv", "q10_pep_used", "q11_pep_want", "q12_prep_heard", "q13_prep_want"]
PAGE7_BASE_KEY = "testing_habit"
PAGE7_TEXT_KEYS = ["phone", "email", "other_contact"]


def make_sample_xlsx(path):
    """v1.0.31：仿衛福部官方雙層樣式 — 標題列深藍底白字 + 描述列淺黃底深字 + 樣本列灰底"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    wb = Workbook()
    ws = wb.active
    ws.title = "匿名HIV取號_範例檔"

    headers = [label for _, label, _, _ in COMPLETE_FIELDS]
    # 描述行（每個欄位的選項提示）
    def _opt_str(allowed, dv):
        if allowed:
            return "/".join(allowed)
        return "（自由文字）"
    desc_row = [_opt_str(allowed, dv) for _, _, dv, allowed in COMPLETE_FIELDS]
    default_row = [dv for _, _, dv, _ in COMPLETE_FIELDS]
    sample = {
        "gender": "女", "year": 1990, "res18": "台南市", "resCur": "台南市",
        "orient": "異性", "edu": "高中職", "testing_habit": "否",
    }
    sample_row = [sample.get(k, dv) for k, _, dv, _ in COMPLETE_FIELDS]

    # Layout：
    #   Row 1 — 標題列（深藍底白字粗體）
    #   Row 2 — 選項提示（淺黃底）
    #   Row 3 — 預設值範本（灰底+斜體，匯入會自動忽略）
    #   Row 4 — 實際資料起點（白色）
    ws.append(headers)
    ws.append(desc_row)
    ws.append(default_row)
    ws.append(sample_row)

    # 樣式
    thin = Side(border_style="thin", color="546E7A")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Row 1 標題：深藍底白字
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=11, name="微軟正黑體")
        c.fill = PatternFill("solid", fgColor="1A237E")  # 深藍
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    # Row 2 描述：淺黃底深字
    for c in ws[2]:
        c.font = Font(color="6D4C41", size=10, name="微軟正黑體")
        c.fill = PatternFill("solid", fgColor="FFF59D")  # 淺黃
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    # Row 3 預設值：灰底+斜體（提示：此列匯入會被自動忽略）
    for c in ws[3]:
        c.font = Font(italic=True, color="78909C", size=9, name="微軟正黑體")
        c.fill = PatternFill("solid", fgColor="ECEFF1")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    # Row 4 範例資料（白底正常）
    for c in ws[4]:
        c.font = Font(color="263238", size=10, name="微軟正黑體")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws["A1"].comment = Comment(
        "規則說明：\n"
        " ‧ 第 1 列：欄位名稱（請勿修改）\n"
        " ‧ 第 2 列：選項提示（請勿修改）\n"
        " ‧ 第 3 列：預設值範本（匯入會自動忽略）\n"
        " ‧ 第 4 列起：填寫實際資料\n\n"
        "提示：每個欄位的儲存格有下拉清單可直接選取，不必手動輸入。",
        "HIV 取號工具"
    )

    # 欄寬：依文字長度估
    for col_idx, label in enumerate(headers, start=1):
        col_letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(16, min(28, len(str(label)) + 4))
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18

    # 下拉資料驗證（從第 4 列開始套用）
    for col_idx, (key, label, dv, allowed) in enumerate(COMPLETE_FIELDS, start=1):
        if not allowed:
            continue
        opts = ",".join([str(v) for v in allowed])
        if len(opts) > 250:
            continue
        col_letter = ws.cell(1, col_idx).column_letter
        dv_obj = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True,
                                showErrorMessage=True,
                                errorTitle="無效選項",
                                error=f"請從下拉選擇：{opts[:60]}...")
        dv_obj.add(f"{col_letter}4:{col_letter}500")
        ws.add_data_validation(dv_obj)
    # 22 縣市改用 named range
    ws_cities = wb.create_sheet("__cities__")
    ws_cities.sheet_state = "hidden"
    for i, c in enumerate(CITIES, start=1):
        ws_cities.cell(i, 1, c)
    cities_range = f"__cities__!$A$1:$A${len(CITIES)}"
    for col_idx, (key, label, dv, allowed) in enumerate(COMPLETE_FIELDS, start=1):
        if allowed is CITIES or (allowed and len(",".join(allowed)) > 250):
            col_letter = ws.cell(1, col_idx).column_letter
            dv_obj = DataValidation(type="list", formula1=f"={cities_range}", allow_blank=True)
            dv_obj.add(f"{col_letter}4:{col_letter}500")
            ws.add_data_validation(dv_obj)

    ws.freeze_panes = "A4"  # 凍結到第 4 列起，前 3 列固定
    wb.save(path)


REQUIRED_KEYS_FOR_WARN = {
    # 一定要填（沒填要警示）
    "gender", "nation", "year", "res18", "resCur", "orient", "edu", "testing_habit",
}

def import_xlsx_profiles(path):
    """v1.0.31/33：讀取 xlsx
       回 (profiles, warnings_dict)
       warnings_dict = {"blank": [...每列未填的欄位...], "invalid": [...無效值...]}"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {"blank": [], "invalid": [], "fatal": "Excel 是空的"}
    header_row = rows[0]
    header_to_key = {label: k for k, label, _, _ in COMPLETE_FIELDS}
    col_keys = []
    for h in header_row:
        h_clean = (str(h or "")).strip()
        col_keys.append(header_to_key.get(h_clean))
    profiles = []
    blanks_by_row = []   # [(row_idx, [field_labels])]
    invalids = []        # [(row_idx, field_label, bad_value, default)]
    # 自動跳過輔助列
    start = 1
    if len(rows) >= 2:
        row2_text = " ".join(str(c or "") for c in rows[1])
        if "/" in row2_text or "自由文字" in row2_text:
            start = max(start, 2)
    if len(rows) >= 3:
        defaults = [dv for _, _, dv, _ in COMPLETE_FIELDS]
        row3 = list(rows[2])
        check_n = min(len(row3), len(defaults))
        if check_n > 0:
            match = sum(1 for i in range(check_n) if str(row3[i] or "") == str(defaults[i] or ""))
            if match >= check_n * 0.6:
                start = max(start, 3)
    for ridx, row in enumerate(rows[start:], start=start + 1):
        if not row or all(c is None or c == "" for c in row):
            continue
        prof_raw = {}
        for i, val in enumerate(row):
            if i >= len(col_keys): break
            k = col_keys[i]
            if k is None: continue
            prof_raw[k] = val
        prof = {}
        row_blanks = []
        for k, _, dv, allowed in COMPLETE_FIELDS:
            actual = prof_raw.get(k)
            is_blank = (actual is None or str(actual).strip() == "")
            if is_blank:
                # 只警示「應填但沒填」的（required 清單），其他靜默套用預設
                if k in REQUIRED_KEYS_FOR_WARN:
                    label = OUTPUT_LABELS.get(k, k)
                    row_blanks.append(label)
                prof[k] = dv
            elif allowed:
                v = str(actual).strip()
                if v not in allowed:
                    label = OUTPUT_LABELS.get(k, k)
                    invalids.append((ridx, label, v, dv))
                    prof[k] = dv
                else:
                    prof[k] = v
            else:
                prof[k] = actual
        if row_blanks:
            blanks_by_row.append((ridx, row_blanks))
        try: prof["year"] = int(prof.get("year") or 1990)
        except Exception: prof["year"] = 1990
        profiles.append(prof)
    return profiles, {"blank": blanks_by_row, "invalid": invalids}


class CanvasProgressBar(tk.Canvas):
    """v1.0.48：Canvas 自繪卡通跑者進度條（取代 v1.0.45 emoji 版）
    參考 Claude Design 概念稿 aRg0bGRxr6wjMMbSFm2vVw exe_concept.html。
    跑者由多個 Canvas 圖元組成，主題切換時更新身體/腿色。
    軌道下方 18px 進度填充 bar，上方放跑者 + 終點旗。
    跑步時上下 ±3px 跳動模擬奔跑。

    注意：`self._w` 與 `self._h` 是 Tk widget 內部屬性，不可覆蓋！用 `_cw`/`_ch`。"""

    SKIN = "#f5c8a3"
    HAIR = "#3a2a1f"
    EYE = "#1a1a1a"
    SHOE = "#2a2a2a"
    POLE = "#4a5567"
    SHADOW = "#bdbdbd"

    def __init__(self, parent, width=600, height=58, bg="#f3f5f8",
                 accent="#2c6bd1", accent_2="#1f4f9c", **kw):
        super().__init__(parent, width=width, height=height, bg=bg,
                         highlightthickness=0, bd=0, **kw)
        self._cw, self._ch = width, height
        self._bg, self._fg = bg, accent
        self._fg_2 = accent_2
        # 底部 18px 進度填充 bar
        bar_top = height - 18
        self._bar_top = bar_top
        self.bar_track = self.create_rectangle(0, bar_top, width, height,
                                                fill="#eef1f5", outline="")
        self.bar = self.create_rectangle(0, bar_top, 0, height, fill=accent, width=0)
        # 跑者基準 y（腳剛好踩在 bar 頂上）
        self._runner_y = bar_top - 6
        # 跑者 cartoon 圖元
        self._runner_items = {}
        self._build_runner(22, self._runner_y)
        # 終點旗
        self._flag_items = {}
        self._build_flag(width - 22, self._runner_y)
        # % 文字（懸浮 bar 上方）
        self.txt = self.create_text(width // 2, bar_top - 2,
                                     text="0.0%", fill="#4a5567",
                                     font=("Consolas", 9, "bold"),
                                     anchor="s")
        self.value = 0
        self.maximum = 100
        self._frame = 0
        self._anim_running = False
        self.bind("<Configure>", self._on_resize)
        self.after(220, self._tick)

    # ── 建構跑者（基準點：腳底中心 (cx, by)）──
    def _build_runner(self, cx, by):
        """畫一隻 24×34 的卡通跑者，腳底中心對齊 (cx, by)
        items 字典記住每個圖元 id，後續可移動/換色"""
        i = self._runner_items
        # 影子（橢圓）
        i['shadow'] = self.create_oval(cx-9, by-1, cx+9, by+3, fill=self.SHADOW, outline="")
        # 後腿（三點折線 smooth）
        i['back_leg'] = self.create_line(cx-2, by-14, cx-6, by-7, cx-9, by-1,
                                          smooth=True, width=3, fill=self._fg_2, capstyle="round")
        # 前腿
        i['front_leg'] = self.create_line(cx+1, by-14, cx+5, by-9, cx+8, by-3,
                                           smooth=True, width=3, fill=self._fg_2, capstyle="round")
        # 鞋
        i['shoe_back'] = self.create_oval(cx-11, by-2, cx-7, by+1, fill=self.SHOE, outline="")
        i['shoe_front'] = self.create_oval(cx+6, by-4, cx+10, by-1, fill=self.SHOE, outline="")
        # 身體（梯形 polygon smooth）— 從腰到肩
        body_pts = [cx-5, by-14, cx+5, by-15, cx+6, by-22, cx-4, by-23]
        i['body'] = self.create_polygon(body_pts, smooth=True, fill=self._fg, outline="")
        # 後手臂
        i['back_arm'] = self.create_line(cx-4, by-22, cx-9, by-20, cx-11, by-17,
                                          smooth=True, width=2.5, fill=self.SKIN, capstyle="round")
        # 前手臂往前甩
        i['front_arm'] = self.create_line(cx+4, by-22, cx+9, by-23, cx+12, by-26,
                                           smooth=True, width=2.5, fill=self.SKIN, capstyle="round")
        # 頭
        i['head'] = self.create_oval(cx-5, by-32, cx+5, by-22, fill=self.SKIN, outline="")
        # 頭髮（從上覆過頭頂的弧）
        hair_pts = [cx-5, by-26, cx-3, by-32, cx+1, by-33, cx+5, by-32, cx+5, by-27,
                    cx+3, by-29, cx, by-28, cx-3, by-29]
        i['hair'] = self.create_polygon(hair_pts, smooth=True, fill=self.HAIR, outline="")
        # 眼睛
        i['eye'] = self.create_oval(cx+1, by-29, cx+2.5, by-27.5, fill=self.EYE, outline="")
        # 嘴角微笑
        i['mouth'] = self.create_line(cx+1, by-25, cx+2.5, by-24.5, cx+4, by-25,
                                       smooth=True, width=1, fill=self.EYE)
        # 汗水（兩顆）
        i['sweat1'] = self.create_oval(cx-9, by-34, cx-7, by-32, fill="#76aef0", outline="")
        i['sweat2'] = self.create_oval(cx-12, by-29, cx-10.5, by-27.5, fill="#76aef0", outline="")
        # 記住基準座標供後續移動/重畫
        self._runner_cx = cx
        self._runner_by = by

    def _build_flag(self, cx, by):
        """終點旗：旗桿 + 三角旗 + 6 顆白格子"""
        f = self._flag_items
        f['pole'] = self.create_line(cx, by-30, cx, by+2, fill=self.POLE, width=2, capstyle="round")
        f['flag'] = self.create_polygon(cx+1, by-30, cx+18, by-30, cx+15, by-25,
                                         cx+18, by-20, cx+1, by-20,
                                         smooth=False, fill=self.POLE, outline="")
        # 6 顆白格子
        f['c1'] = self.create_rectangle(cx+3, by-29, cx+5, by-27, fill="white", outline="")
        f['c2'] = self.create_rectangle(cx+9, by-29, cx+11, by-27, fill="white", outline="")
        f['c3'] = self.create_rectangle(cx+6, by-26, cx+8, by-24, fill="white", outline="")
        f['c4'] = self.create_rectangle(cx+12, by-26, cx+14, by-24, fill="white", outline="")
        f['c5'] = self.create_rectangle(cx+3, by-23, cx+5, by-21, fill="white", outline="")
        f['c6'] = self.create_rectangle(cx+9, by-23, cx+11, by-21, fill="white", outline="")
        self._flag_cx = cx
        self._flag_by = by

    def _move_runner_to(self, new_cx, new_by):
        """整組跑者圖元位移到新基準點"""
        dx = new_cx - self._runner_cx
        dy = new_by - self._runner_by
        if dx == 0 and dy == 0:
            return
        for item_id in self._runner_items.values():
            self.move(item_id, dx, dy)
        self._runner_cx = new_cx
        self._runner_by = new_by

    def _move_flag_to(self, new_cx):
        """旗子只 x 方向（resize 時）"""
        dx = new_cx - self._flag_cx
        if dx == 0:
            return
        for item_id in self._flag_items.values():
            self.move(item_id, dx, 0)
        self._flag_cx = new_cx

    def _tick(self):
        # 奔跑中：上下 ±3px 跳動
        if self._anim_running:
            self._frame = 1 - self._frame
            new_by = self._runner_y + (-3 if self._frame else 0)
            self._move_runner_to(self._runner_cx, new_by)
        try:
            self.after(220, self._tick)
        except Exception:
            pass

    def _on_resize(self, event):
        self._cw = event.width
        self.coords(self.bar_track, 0, self._bar_top, self._cw, self._ch)
        self.coords(self.txt, self._cw // 2, self._bar_top - 2)
        self._move_flag_to(self._cw - 22)
        self._refresh()

    def configure_value(self, value, maximum=None):
        if maximum is not None:
            self.maximum = max(1, maximum)
        self.value = max(0, min(value, self.maximum))
        pct = self.value / self.maximum if self.maximum else 0
        self._anim_running = 0 < pct < 1
        if not self._anim_running:
            self._move_runner_to(self._runner_cx, self._runner_y)
            self._frame = 0
        self._refresh()

    def _refresh(self):
        pct = (self.value / self.maximum) if self.maximum else 0
        bar_w = int(self._cw * pct)
        self.coords(self.bar, 0, self._bar_top, bar_w, self._ch)
        self.itemconfigure(self.txt, text=f"{pct*100:.1f}%")
        # 跑者 x：起點 22 → 右邊 _cw - 50（避開旗子）
        x_min, x_max = 22, max(22, self._cw - 50)
        runner_cx = int(x_min + (x_max - x_min) * pct)
        # 保留當前 y（可能在跳動中）
        cur_by = self._runner_by
        self._move_runner_to(runner_cx, cur_by)

    def configure_theme(self, bg, fg, fg_2=None):
        """主題切換：bg、進度條色、跑者衣服色（runner-2 用 fg_2 或 fg 暗化）"""
        self._bg, self._fg = bg, fg
        if fg_2:
            self._fg_2 = fg_2
        self.configure(bg=bg)
        self.itemconfigure(self.bar, fill=fg)
        self.itemconfigure(self._runner_items['body'], fill=fg)
        self.itemconfigure(self._runner_items['back_leg'], fill=self._fg_2)
        self.itemconfigure(self._runner_items['front_leg'], fill=self._fg_2)

    # ── v1.0.51 跑者狀態（Claude Design 概念稿 3.6）──
    # state: 'idle' | 'running' | 'arrived' | 'panting'
    # idle  → 清空 overlay，靜止
    # running → 已由 configure_value 自動處理（跳動）
    # arrived → 🎉 + confetti 1.5s 後自動回 idle
    # panting → 💦 持續顯示，下次 set_state 才清
    def set_state(self, state):
        # 先清掉之前的 overlay
        if not hasattr(self, '_overlay_items'):
            self._overlay_items = []
        for item in self._overlay_items:
            try: self.delete(item)
            except Exception: pass
        self._overlay_items = []
        if not hasattr(self, '_arrival_after'):
            self._arrival_after = None
        if self._arrival_after:
            try: self.after_cancel(self._arrival_after)
            except Exception: pass
            self._arrival_after = None

        if state == 'arrived':
            cx, by = self._runner_cx, self._runner_by
            # 🎉 emoji 在跑者頭上方
            tid = self.create_text(cx, by - 38, text="🎉",
                                    font=("Segoe UI Emoji", 14))
            self._overlay_items.append(tid)
            # 撒花：8 顆小色塊散落跑者周圍
            import random
            colors = ["#5b8def", "#d97a8a", "#5fa787", "#c39145", "#6c5ce7", "#0fa3a3"]
            for _ in range(8):
                ox = random.randint(-30, 30)
                oy = random.randint(-32, -8)
                cc = random.choice(colors)
                rx, ry = cx + ox, by + oy
                rid = self.create_rectangle(rx, ry, rx + 3, ry + 4,
                                             fill=cc, outline="")
                self._overlay_items.append(rid)
            # 1.5 秒後自動回 idle
            self._arrival_after = self.after(1500, lambda: self.set_state('idle'))
        elif state == 'panting':
            cx, by = self._runner_cx, self._runner_by
            tid = self.create_text(cx + 10, by - 38, text="💦",
                                    font=("Segoe UI Emoji", 12))
            self._overlay_items.append(tid)
        # idle / running 不需 overlay（已清掉）

# ── 工具函式 ──────────────────────────────────────
def _hide_path(path):
    """v1.0.40/42：把 Windows 檔案或資料夾標 HIDDEN 屬性，工作目錄不要長一堆礙眼"""
    try:
        import ctypes
        FILE_ATTRIBUTE_HIDDEN = 0x02
        ctypes.windll.kernel32.SetFileAttributesW(str(path), FILE_ATTRIBUTE_HIDDEN)
    except Exception:
        pass


def ensure_outdir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for sub in ("old", "log", "debug"):
        p = os.path.join(OUTPUT_DIR, sub)
        os.makedirs(p, exist_ok=True)
        _hide_path(p)
    # v1.0.53：啟動時掃一次，把其他輔助檔藏起來（不留東西在 EXE 旁邊礙眼）
    #           僅 EXE + 諮詢代碼.xlsx 是「使用者要看到的」
    for fname in ("settings.json", "resume_state.json"):
        p = os.path.join(OUTPUT_DIR, fname)
        if os.path.isfile(p):
            _hide_path(p)
    if os.path.isfile(SETTINGS_FILE):
        _hide_path(SETTINGS_FILE)
    # 舊版（v1.0.52 之前）的 *_HIV_CODE.xlsx 累積在主目錄會礙眼
    # 不直接搬（資料寶貴），僅在沒「諮詢代碼.xlsx」時 archive_old_outputs 會搬

# ── 全域 log file（每次啟動 GUI 開新檔） ──
_LOG_FP = None
def init_logfile():
    global _LOG_FP
    ensure_outdir()
    fp = os.path.join(OUTPUT_DIR, "log",
        f"log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    try:
        _LOG_FP = open(fp, "w", encoding="utf-8", buffering=1)  # line-buffered
    except Exception:
        _LOG_FP = None
    return fp

def write_log_line(line):
    if _LOG_FP:
        try:
            _LOG_FP.write(line + "\n")
        except Exception:
            pass

def debug_dir():
    p = os.path.join(OUTPUT_DIR, "debug")
    os.makedirs(p, exist_ok=True)
    return p

def archive_old_outputs(log_fn=None):
    """把 number/ 下「非今天」的 *.xlsx / *.csv 搬到 number/old/。
       v1.0.17：保留今天的檔案以便同日合併（merge 邏輯在 _incremental_save_excel）
       v1.0.53：新檔名「諮詢代碼.xlsx」是固定持久檔，多日多分頁；不再 archive 它"""
    old_dir = os.path.join(OUTPUT_DIR, "old")
    os.makedirs(old_dir, exist_ok=True)
    today = datetime.datetime.now().strftime("%Y%m%d")
    moved = 0
    try:
        for fn in os.listdir(OUTPUT_DIR):
            full = os.path.join(OUTPUT_DIR, fn)
            if not os.path.isfile(full):
                continue
            if not (fn.lower().endswith(".xlsx") or fn.lower().endswith(".csv")):
                continue
            # v1.0.36：今天的 HIV_CODE 檔留下供同日合併
            if fn == f"{today}_HIV_CODE.xlsx":
                continue
            # v1.0.53：固定檔名「諮詢代碼.xlsx」永久保留（多日多分頁累積）
            if fn == "諮詢代碼.xlsx" or fn.startswith("諮詢代碼") and fn.endswith(".xlsx"):
                continue
            try:
                import shutil
                target = os.path.join(old_dir, fn)
                if os.path.exists(target):
                    stem, ext = os.path.splitext(fn)
                    target = os.path.join(old_dir,
                        f"{stem}_{datetime.datetime.now().strftime('%H%M%S')}{ext}")
                shutil.move(full, target)
                moved += 1
                if log_fn: log_fn(f"  → 舊檔搬移：{fn}")
            except Exception as e:
                if log_fn: log_fn(f"  ⚠ 搬移失敗 {fn}：{e}")
    except FileNotFoundError:
        pass
    if moved and log_fn:
        log_fn(f"📦 已將 {moved} 個舊檔搬至 number\\old\\")
    return moved

def save_debug_snapshot(driver, tag):
    """DEBUG 版：失敗時把當下頁面 HTML + screenshot 存到 number/debug/"""
    if not DEBUG or not driver:
        return
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
        base = os.path.join(debug_dir(), f"{ts}_{tag}")
        with open(base + ".html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        try: driver.save_screenshot(base + ".png")
        except Exception: pass
    except Exception:
        pass

def weighted_pick(items, weights):
    """依比例權重隨機抽一個項目（weights 不必正規化）"""
    if not items:
        return None
    total = sum(max(0, w) for w in weights)
    if total <= 0:
        return random.choice(items)
    r = random.uniform(0, total)
    acc = 0
    for it, w in zip(items, weights):
        acc += max(0, w)
        if r <= acc:
            return it
    return items[-1]

def jitter(a=0.6, b=1.6):
    """模擬人類延遲（純函式版本，作為 fallback）"""
    time.sleep(random.uniform(a, b))


class DelayConfig:
    """從 GUI 拿到的延遲設定，由 HivaWorker 引用"""
    def __init__(self, action_lo=0.3, action_hi=0.7,
                 page_lo=0.8, page_hi=1.5,
                 between_lo=1.5, between_hi=3.0,
                 wait_timeout=30):
        self.action_lo  = float(action_lo)
        self.action_hi  = float(action_hi)
        self.page_lo    = float(page_lo)
        self.page_hi    = float(page_hi)
        self.between_lo = float(between_lo)
        self.between_hi = float(between_hi)
        self.wait_timeout = int(wait_timeout)  # selenium 等待逾時（多人連線會慢，預設拉長）

    def action(self):  time.sleep(random.uniform(self.action_lo, self.action_hi))
    def page(self):    time.sleep(random.uniform(self.page_lo,   self.page_hi))
    def between(self): time.sleep(random.uniform(self.between_lo, self.between_hi))


# ── Selenium：取代碼工人 ──────────────────────────
class HivaWorker:
    def __init__(self, log_fn, status_fn, code_fn, stop_evt, delay_cfg=None):
        self.log = log_fn
        self.status = status_fn
        self.on_code = code_fn
        self.stop_evt = stop_evt
        self.driver = None
        self.dly = delay_cfg or DelayConfig()

    def _import_selenium(self):
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options
        from selenium.webdriver.edge.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.webdriver.support import expected_conditions as EC
        return webdriver, Options, Service, By, WebDriverWait, Select, EC

    def start_browser(self):
        webdriver, Options, Service, By, WebDriverWait, Select, EC = self._import_selenium()
        opts = Options()
        # ── 啟動參數：移除自動化痕跡 ──
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "load-extension"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument("--start-maximized")
        opts.add_argument("--lang=zh-TW")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--disable-notifications")
        opts.add_argument("--disable-popup-blocking")
        opts.add_argument("--disable-features=IsolateOrigins,site-per-process")
        opts.add_argument("--disable-site-isolation-trials")
        # 真實 UA（與 Edge 130 相符）
        opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0")
        # prefs：關掉密碼儲存提示等
        opts.add_experimental_option("prefs", {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.notifications": 2,
        })
        try:
            self.driver = webdriver.Edge(options=opts)
        except Exception as e:
            self.log(f"✗ 啟動 Edge 失敗：{e}")
            self.log("   請先安裝 Edge WebDriver（https://developer.microsoft.com/microsoft-edge/tools/webdriver/）")
            return False

        # ── stealth：25+ 項偽裝（覆蓋常見指紋偵測手法） ──
        stealth_js = r"""
        // 1. webdriver
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        delete Object.getPrototypeOf(navigator).webdriver;

        // 2. languages
        Object.defineProperty(navigator, 'languages', {get: () => ['zh-TW','zh','en-US','en']});

        // 3. plugins（模擬 5 個 plugin，PDF Viewer 等）
        Object.defineProperty(navigator, 'plugins', {get: () => {
            const p = [
                {name: 'PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format'},
                {name: 'Chrome PDF Viewer', filename: 'internal-pdf-viewer', description: ''},
                {name: 'Chromium PDF Viewer', filename: 'internal-pdf-viewer', description: ''},
                {name: 'Microsoft Edge PDF Viewer', filename: 'internal-pdf-viewer', description: ''},
                {name: 'WebKit built-in PDF', filename: 'internal-pdf-viewer', description: ''},
            ];
            p.item = i => p[i]; p.namedItem = n => p.find(x=>x.name===n); p.refresh = ()=>{};
            return p;
        }});

        // 4. mimeTypes
        Object.defineProperty(navigator, 'mimeTypes', {get: () => [
            {type: 'application/pdf', suffixes: 'pdf', description: ''},
            {type: 'text/pdf', suffixes: 'pdf', description: ''}
        ]});

        // 5. chrome 物件
        if (!window.chrome) window.chrome = {};
        window.chrome.runtime = window.chrome.runtime || { OnInstalledReason:{}, OnRestartRequiredReason:{}, PlatformOs:{}, PlatformArch:{}, RequestUpdateCheckStatus:{} };
        window.chrome.app = window.chrome.app || { isInstalled: false, InstallState:{}, RunningState:{} };
        window.chrome.csi = window.chrome.csi || function() { return { startE: Date.now(), onloadT: Date.now(), pageT: Date.now(), tran: 15 }; };
        window.chrome.loadTimes = window.chrome.loadTimes || function() { return { requestTime: Date.now()/1000, startLoadTime: Date.now()/1000, commitLoadTime: Date.now()/1000, finishDocumentLoadTime: Date.now()/1000, finishLoadTime: Date.now()/1000, firstPaintTime: Date.now()/1000, firstPaintAfterLoadTime: 0, navigationType: 'Other', wasFetchedViaSpdy: false, wasNpnNegotiated: false, npnNegotiatedProtocol: 'unknown', wasAlternateProtocolAvailable: false, connectionInfo: 'http/1.1' }; };

        // 6. permissions query mock（notifications 修正常見偵測點）
        const origQuery = window.navigator.permissions && window.navigator.permissions.query;
        if (origQuery) {
            window.navigator.permissions.query = (p) => p.name === 'notifications'
                ? Promise.resolve({ state: Notification.permission })
                : origQuery(p);
        }

        // 7~10. screen / hardware
        Object.defineProperty(screen, 'colorDepth', {get: () => 24});
        Object.defineProperty(screen, 'pixelDepth', {get: () => 24});
        Object.defineProperty(navigator, 'platform', {get: () => 'Win32'});
        Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
        Object.defineProperty(navigator, 'deviceMemory', {get: () => 8});

        // 11. userAgentData（Edge 新版指紋）
        Object.defineProperty(navigator, 'userAgentData', {get: () => ({
            brands: [
                {brand: 'Microsoft Edge', version: '130'},
                {brand: 'Chromium', version: '130'},
                {brand: 'Not?A_Brand', version: '99'},
            ],
            mobile: false,
            platform: 'Windows',
            getHighEntropyValues: () => Promise.resolve({
                architecture: 'x86', bitness: '64', model: '',
                platformVersion: '15.0.0', uaFullVersion: '130.0.0.0',
                fullVersionList: [{brand:'Microsoft Edge',version:'130.0.0.0'},{brand:'Chromium',version:'130.0.0.0'}]
            })
        })});

        // 12. WebGL vendor / renderer 偽裝
        const getParameter = WebGLRenderingContext.prototype.getParameter;
        WebGLRenderingContext.prototype.getParameter = function(p) {
            if (p === 37445) return 'Intel Inc.';
            if (p === 37446) return 'Intel Iris OpenGL Engine';
            return getParameter.call(this, p);
        };
        if (window.WebGL2RenderingContext) {
            const getParameter2 = WebGL2RenderingContext.prototype.getParameter;
            WebGL2RenderingContext.prototype.getParameter = function(p) {
                if (p === 37445) return 'Intel Inc.';
                if (p === 37446) return 'Intel Iris OpenGL Engine';
                return getParameter2.call(this, p);
            };
        }

        // 13. Canvas 微擾動（避免 fingerprint hash 命中黑名單）
        const toDataURL = HTMLCanvasElement.prototype.toDataURL;
        HTMLCanvasElement.prototype.toDataURL = function(...a) {
            const ctx = this.getContext('2d');
            if (ctx) { ctx.fillStyle = 'rgba(0,0,0,0.005)'; ctx.fillRect(0,0,1,1); }
            return toDataURL.apply(this, a);
        };

        // 14. battery API mock（很多偵測會問）
        if (navigator.getBattery) {
            navigator.getBattery = () => Promise.resolve({
                charging: true, chargingTime: 0, dischargingTime: Infinity, level: 1,
                addEventListener: ()=>{}, removeEventListener: ()=>{}, dispatchEvent: ()=>true
            });
        }

        // 15. connection
        Object.defineProperty(navigator, 'connection', {get: () => ({
            effectiveType: '4g', rtt: 50, downlink: 10, saveData: false,
            addEventListener: ()=>{}, removeEventListener: ()=>{}
        })});

        // 16. mediaDevices
        if (navigator.mediaDevices) {
            const enumOrig = navigator.mediaDevices.enumerateDevices;
            navigator.mediaDevices.enumerateDevices = () => Promise.resolve([
                {deviceId:'default', kind:'audioinput', label:'', groupId:''},
                {deviceId:'default', kind:'audiooutput', label:'', groupId:''},
            ]);
        }

        // 17. AudioContext 微擾（指紋擾動）
        if (window.AudioContext || window.webkitAudioContext) {
            const AC = window.AudioContext || window.webkitAudioContext;
            const orig = AC.prototype.createOscillator;
            AC.prototype.createOscillator = function() {
                const osc = orig.apply(this, arguments);
                const f = osc.frequency.value;
                osc.frequency.value = f + Math.random() * 0.0001;
                return osc;
            };
        }

        // 18. cookieEnabled / doNotTrack
        Object.defineProperty(navigator, 'cookieEnabled', {get: () => true});
        Object.defineProperty(navigator, 'doNotTrack', {get: () => null});

        // 19. webdriver in iframes（contentWindow 也要清）
        const _addEventListener = document.addEventListener;
        document.addEventListener = function(t, fn, ...rest) {
            if (t === 'DOMContentLoaded') {
                document.querySelectorAll('iframe').forEach(f => {
                    try {
                        Object.defineProperty(f.contentWindow.navigator, 'webdriver', {get:()=>undefined});
                    } catch(_){}
                });
            }
            return _addEventListener.call(this, t, fn, ...rest);
        };

        // 20. 隱藏 Function.prototype.toString 對 native function 的洩漏
        const nativeToString = Function.prototype.toString;
        Function.prototype.toString = function() {
            if (this === Function.prototype.toString) return nativeToString.call(this);
            return nativeToString.call(this);
        };

        // 21. Notification.permission（避免 'denied' 露餡）
        if (window.Notification) {
            Object.defineProperty(Notification, 'permission', {get: () => 'default'});
        }

        // 22. window.outerWidth/Height 補正（無頭模式露餡防呆）
        if (window.outerWidth === 0) {
            Object.defineProperty(window, 'outerWidth', {get: () => window.innerWidth});
            Object.defineProperty(window, 'outerHeight', {get: () => window.innerHeight});
        }

        // 23. 移除 cdc_*/$cdc_* 變數（webdriver 洩漏點）
        for (const k of Object.keys(window)) {
            if (/^[$_]?cdc_/.test(k) || /^[$_]?wdc_/.test(k)) {
                try { delete window[k]; } catch(_){}
            }
        }

        // 24. CSS prefers-color-scheme 等媒體查詢
        // 保留瀏覽器預設

        // 25. 移除 navigator.bluetooth（automation 通常沒有）
        try { delete Navigator.prototype.bluetooth; } catch(_){}
        """
        try:
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": stealth_js})
        except Exception:
            pass
        # 額外：CDP 改 User-Agent metadata
        try:
            self.driver.execute_cdp_cmd("Network.setUserAgentOverride", {
                "userAgent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                             "(KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0",
                "acceptLanguage": "zh-TW,zh;q=0.9,en;q=0.8",
                "platform": "Win32",
            })
        except Exception:
            pass
        return True

    def quit(self):
        if self.driver:
            try: self.driver.quit()
            except Exception: pass
            self.driver = None

    # ── 流程：取一筆 ──────────────────────────────
    def fetch_one(self, profile):
        """profile: dict — gender/nation/year/res18/resCur/orient/edu/testing"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.webdriver.support import expected_conditions as EC
        d = self.driver
        wait = WebDriverWait(d, self.dly.wait_timeout)

        # 1) 首頁 → 點「進入風險評估」
        d.get(URL_HOME)
        self.dly.page()
        try:
            link = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(),'進入風險評估') or contains(text(),'風險評估')]")))
            try: link.click()
            except Exception:
                d.execute_script("arguments[0].click();", link)
        except Exception as e:
            self.log(f"⚠ 找不到「進入風險評估」按鈕：{e}")
            save_debug_snapshot(d, "home_no_risk_btn")
            return None
        # 等到 Page 1 載入（只等 QuestWizard radio 出現；URL 條件已證實會誤判）
        try:
            WebDriverWait(d, self.dly.wait_timeout).until(
                lambda drv: len(drv.find_elements(By.XPATH,
                    "//input[@type='radio' and contains(@name,'QuestWizard')]")) > 0
            )
        except Exception:
            self.log("⚠ 進入問卷後，Page 1 沒等到 radio，續行嘗試")
        self.dly.action()

        # v1.0.22：偵測完整模式 — profile 含 q1_sex 即為完整 profile
        is_complete = "q1_sex" in profile

        # === Page 1 ===
        # v1.0.24：Q1=否 → 系統自動帶 Q2-5 = 沒有發生，只點 Q1 即可
        # Q1=是 才需要逐一填 Q2-Q5
        if is_complete:
            if profile.get("q1_sex") == "否":
                # 只點 Q1，避免動到系統自動填的 Q2-5
                p1_keys = [PAGE1_KEYS[0]]
            else:
                p1_keys = PAGE1_KEYS  # Q1=是 → 完整填 Q1-Q5
            n = self._fill_groups_in_order(profile, p1_keys)
            if n == 0:
                self.log("✗ Page 1 沒填到任何 radio"); save_debug_snapshot(d, "p1_fail"); return None
            recovery_p1 = lambda: self._fill_groups_in_order(profile, p1_keys)
        else:
            if not self._click_first_group_no():
                self.log("✗ Page 1 Q1 失敗"); save_debug_snapshot(d, "p1_q1_fail"); return None
            recovery_p1 = lambda: self._click_first_group_no()
        self.dly.action()
        if not self._click_next(expect_progress=15, recovery_fn=recovery_p1):
            save_debug_snapshot(d, "p1_to_p2_fail"); return None
        self.dly.page()

        # === Page 2 ===
        if is_complete:
            self._fill_groups_in_order(profile, [PAGE2_BASE_KEY])
            # 若 Q6=是 → 等子題出現再填
            if profile.get("q6_std") == "是":
                self.log("  ⏳ Q6=是，等子題展開…")
                from selenium.webdriver.support.ui import WebDriverWait
                try:
                    WebDriverWait(d, self.dly.wait_timeout).until(
                        lambda drv: len(self._get_questwizard_groups()) >= 1 + len(PAGE2_SUB_KEYS) - 2
                    )
                    self._fill_groups_in_order(profile, PAGE2_SUB_KEYS, start_index=1)
                    # text 欄位
                    self._fill_text_input_by_label("原因", profile.get("q6_hiv_reason", ""))
                    self._fill_text_input_by_label("其他", profile.get("q6_other", ""))
                except Exception as e:
                    self.log(f"  ⚠ Q6 子題填寫失敗：{e}")
            recovery_p2 = lambda: self._fill_groups_in_order(profile, [PAGE2_BASE_KEY])
        else:
            if not self._click_first_group_no():
                self.log("✗ Page 2 Q6 失敗"); save_debug_snapshot(d, "p2_q6_fail"); return None
            recovery_p2 = lambda: self._click_first_group_no()
        self.dly.action()
        if not self._click_next(expect_progress=30, recovery_fn=recovery_p2):
            save_debug_snapshot(d, "p2_to_p3_fail"); return None
        self.dly.page()

        # === Page 3 ===
        if is_complete:
            self._fill_groups_in_order(profile, [PAGE3_BASE_KEY])
            if profile.get("q7_drug_use") == "是":
                self.log("  ⏳ Q7=是，等子題展開…")
                from selenium.webdriver.support.ui import WebDriverWait
                try:
                    WebDriverWait(d, self.dly.wait_timeout).until(
                        lambda drv: len(self._get_questwizard_groups()) >= 1 + len(PAGE3_SUB_KEYS) - 2
                    )
                    self._fill_groups_in_order(profile, PAGE3_SUB_KEYS, start_index=1)
                    self._fill_text_input_by_label("其他", profile.get("q7_other", ""))
                    # v1.0.23 安非他命使用方式：多 checkbox（吸入/注射/口服）
                    if profile.get("q7_amph") == "是":
                        method_str = profile.get("q7_amph_method", "") or ""
                        methods = [s.strip() for s in str(method_str).replace("，", ",").split(",") if s.strip()]
                        if methods:
                            n = self._click_checkboxes_by_text(methods)
                            self.log(f"  ✓ 安非他命使用方式 勾選 {n}/{len(methods)} 項")
                except Exception as e:
                    self.log(f"  ⚠ Q7 子題填寫失敗：{e}")
            recovery_p3 = lambda: self._fill_groups_in_order(profile, [PAGE3_BASE_KEY])
        else:
            if not self._click_first_group_no():
                self.log("✗ Page 3 Q7 失敗"); save_debug_snapshot(d, "p3_q7_fail"); return None
            recovery_p3 = lambda: self._click_first_group_no()
        self.dly.action()
        if not self._click_next(expect_progress=45, recovery_fn=recovery_p3):
            save_debug_snapshot(d, "p3_to_p4_fail"); return None
        self.dly.page()

        # === Page 4 ===
        if is_complete:
            n = self._fill_groups_in_order(profile, PAGE4_KEYS)
            if n < 4:
                self.log(f"⚠ Page 4 只填到 {n}/4");
            recovery_p4 = lambda: self._fill_groups_in_order(profile, PAGE4_KEYS)
        else:
            if not self._click_all_groups_no(expected_n=4):
                self.log("✗ Page 4 Q8 失敗"); save_debug_snapshot(d, "p4_q8_fail"); return None
            recovery_p4 = lambda: self._click_all_groups_no(expected_n=4)
        if not self._click_next(expect_progress=60, recovery_fn=recovery_p4):
            save_debug_snapshot(d, "p4_to_p5_fail"); return None
        self.dly.page()

        # === Page 5 ===
        if is_complete:
            n = self._fill_groups_in_order(profile, PAGE5_KEYS)
            if n < 5:
                self.log(f"⚠ Page 5 只填到 {n}/5")
            recovery_p5 = lambda: self._fill_groups_in_order(profile, PAGE5_KEYS)
        else:
            if not self._click_all_groups_no(expected_n=5):
                self.log("✗ Page 5 Q9~13 失敗"); save_debug_snapshot(d, "p5_q9to13_fail"); return None
            recovery_p5 = lambda: self._click_all_groups_no(expected_n=5)
        if not self._click_next(expect_progress=75,
                                expect_element_id="ctl00_MainContent_QuestWizard_BirthYear",
                                recovery_fn=recovery_p5):
            save_debug_snapshot(d, "p5_to_p6_fail"); return None
        self.dly.page()

        # === Page 6 — 基本資料（v1.0.32：實際 radio name 為「心理性別」非「性別」） ===
        gender_map = {"男": "1", "女": "2", "跨性別": "3", "其他": "4"}
        nation_map = {"本國籍": "1", "外國籍": "2"}
        if not self._set_radio_by_name("ctl00$MainContent$QuestWizard$心理性別",
                                        gender_map.get(profile["gender"], "1")):
            self._click_label_text(profile["gender"])
        self.dly.action()
        if not self._set_radio_by_name("ctl00$MainContent$QuestWizard$國籍",
                                        nation_map.get(profile["nation"], "1")):
            self._click_label_text(profile["nation"])
        self.dly.action()
        # 6-2 出生年（ASP.NET PostBack 控件，選了會回 server 重畫頁面）
        if not self._select_by_id("ctl00_MainContent_QuestWizard_BirthYear", str(profile["year"])):
            self.log("✗ Page 6 出生年下拉失敗"); save_debug_snapshot(d, "p6_year_fail"); return None
        # 等 PostBack 完成 — 18歲前居住地會從 disabled 變 enabled
        self.log(f"  ⏳ 出生年 {profile['year']} 已選，等 PostBack…")
        if not self._wait_for_enabled("ctl00_MainContent_QuestWizard_DDL_Live_18_ago"):
            save_debug_snapshot(d, "p6_postback_timeout"); return None
        self.dly.action()
        # 6-3 18 歲以前居住地
        if not self._select_by_id("ctl00_MainContent_QuestWizard_DDL_Live_18_ago", profile["res18"]):
            self.log("✗ Page 6 18歲前居住地失敗"); save_debug_snapshot(d, "p6_res18_fail"); return None
        self.dly.action()
        # 6-4 現居住地（id 是「居住地」中文）
        if not self._select_by_id("ctl00_MainContent_QuestWizard_居住地", profile["resCur"]):
            self.log("✗ Page 6 現居住地失敗"); save_debug_snapshot(d, "p6_resCur_fail"); return None
        self.dly.action()
        # 6-5 性傾向（name + value：1=同 2=雙 3=異）
        orient_map = {"同性": "1", "雙性": "2", "異性": "3"}
        if not self._set_radio_by_name("ctl00$MainContent$QuestWizard$性傾向",
                                        orient_map.get(profile["orient"], "3")):
            save_debug_snapshot(d, "p6_orient_fail"); return None
        self.dly.action()
        # 6-6 教育程度（name + value：1=不識字 2=國中以下 3=高中職 4=專科或大學 5=研究所）
        edu_map = {"不識字": "1", "國中以下": "2", "高中職": "3", "專科或大學": "4", "研究所(含)以上": "5"}
        if not self._set_radio_by_name("ctl00$MainContent$QuestWizard$教育程度",
                                        edu_map.get(profile["edu"], "3")):
            save_debug_snapshot(d, "p6_edu_fail"); return None
        self.dly.action()
        if not self._click_next(expect_progress=90):
            self._dismiss_alert_if_any(); save_debug_snapshot(d, "p6_to_p7_fail"); return None
        self.dly.page()

        # === Page 7 — 篩檢習慣 + 聯絡方式（v1.0.32：用實際 name + value） ===
        if is_complete:
            # testing_habit (rdlRegularScreening): 1=是 / 2=否 / 0=從未做過
            th_map = {"是": "1", "否": "2", "從未做過": "0"}
            tval = th_map.get(profile.get("testing_habit", "否"), "2")
            if not self._set_radio_by_name("ctl00$MainContent$QuestWizard$rdlRegularScreening", tval):
                # fallback：原 fill_groups
                self._fill_groups_in_order(profile, [PAGE7_BASE_KEY])
            # text 欄位用 name 精確定位（v1.0.32 修正：之前 label 鄰近搜尋抓不到）
            for key, name in [("phone",         "ctl00$MainContent$QuestWizard$txtMobile"),
                               ("email",         "ctl00$MainContent$QuestWizard$txtEMail"),
                               ("other_contact", "ctl00$MainContent$QuestWizard$txtOther")]:
                v = profile.get(key, "")
                if v:
                    self._fill_text_input_by_name(name, v)
        else:
            choice = profile.get("testing", "否")
            if choice == "從未做過":
                self._click_label_text("從未做過")
            else:
                self._click_label_text("否")
        self.dly.action()
        if not self._click_done():
            return None
        self.dly.page()

        # === 結果頁：抓代碼（伺服器繁忙時可能要等較久） ===
        try:
            WebDriverWait(d, self.dly.wait_timeout * 2).until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'諮詢代碼')]")))
        except Exception:
            self.log("⚠ 沒等到結果頁")
            save_debug_snapshot(d, "result_page_timeout")
            return None
        # v1.0.12：DEBUG 模式存前 2 筆成功的結果頁，方便驗證代碼讀對位置
        if DEBUG and getattr(self, "_result_sample_count", 0) < 2:
            save_debug_snapshot(d, f"result_page_sample_{getattr(self, '_result_sample_count', 0) + 1}")
            self._result_sample_count = getattr(self, "_result_sample_count", 0) + 1
        # v1.0.12：用 body.text 抽（避開 HTML noise / 隱藏 script 文字）
        import re
        try:
            body_text = d.find_element(By.TAG_NAME, "body").text
        except Exception:
            body_text = d.page_source
        # A) 「諮詢代碼為"XXXXXX"」最常見格式
        m = re.search(r'諮詢代碼為["\'“”「]?\s*([0-9A-Z\-]{4,})', body_text)
        # B) 「諮詢代碼: XXXXXX」備用格式
        if not m:
            m = re.search(r'諮詢代碼[\s::是為]+["\'“”「]?\s*([0-9A-Z\-]{4,})', body_text)
        # C) 找「諮詢代碼」之後 30 字內的數字串
        if not m:
            m = re.search(r'諮詢代碼[^0-9]{0,30}([0-9]{4,})', body_text)
        if not m:
            self.log("⚠ 無法解析諮詢代碼，body.text 開頭：" + body_text[:200].replace("\n", " | "))
            save_debug_snapshot(d, "code_not_found")
            return None
        return m.group(1)

    # ── DOM helpers ──────────────────────────────
    def _click_next_btn(self):
        """純粹按下一步按鈕，不做進度驗證"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        d = self.driver
        try:
            btn = WebDriverWait(d, self.dly.wait_timeout).until(EC.element_to_be_clickable(
                (By.XPATH, "//*[(self::button or self::a or self::input) and (contains(.,'下一步') or @value='下一步')]")))
            try: btn.click()
            except Exception: d.execute_script("arguments[0].click();", btn)
            return True
        except Exception as e:
            self.log(f"✗ 找不到下一步按鈕：{e}")
            return False

    def _click_next(self, expect_progress=None, expect_element_id=None,
                    timeout=None, recovery_fn=None, max_retries=2):
        """v1.0.13：雙錨點等待 — 完成度% AND 目標元素 ID 都要到位才算換頁成功。
           單看完成度會被誤導：伺服器有時驗證失敗仍把 % 文字提前更新。
           失敗時最多重試 max_retries 次（每次都呼叫 recovery_fn 補填）。"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        d = self.driver
        to = timeout if timeout is not None else self.dly.wait_timeout

        for attempt in range(max_retries + 1):
            if not self._click_next_btn():
                return False
            ok = True
            if expect_progress is not None:
                ok = self._wait_for_progress(expect_progress, timeout=to)
            if ok and expect_element_id is not None:
                # 二次驗證：目標元素是否真的出現（防進度文字提前但畫面沒切的偽陽性）
                try:
                    WebDriverWait(d, 8).until(EC.presence_of_element_located(
                        (By.ID, expect_element_id)))
                except Exception:
                    self.log(f"⚠ 完成度到了但找不到 #{expect_element_id} — 換頁假性成功")
                    ok = False
            if ok:
                if attempt > 0:
                    self.log(f"✓ 第 {attempt+1} 次嘗試成功")
                return True
            # ── 失敗：偵測+關閉彈窗、補填、重試 ──
            self.log(f"⚠ 換頁失敗（第 {attempt+1} 次），啟動恢復…")
            self._dismiss_alert_if_any()
            time.sleep(0.5)
            if recovery_fn:
                try:
                    self.log("  ↻ 重新填寫本頁缺漏…")
                    recovery_fn()
                except Exception as e:
                    self.log(f"  ↻ recovery_fn 失敗：{e}")
            time.sleep(0.3)
        self.log(f"✗ 重試 {max_retries} 次仍無法換頁，放棄本筆")
        return False

    def _wait_for_progress(self, target_pct, timeout=None):
        """等到頁面顯示「完成度：X %」=target_pct（精確錨點頁面切換）
           多人連線時頁面回應會慢，timeout 從 GUI 設定（預設 30s）"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        d = self.driver
        to = timeout if timeout is not None else self.dly.wait_timeout
        try:
            WebDriverWait(d, to).until(
                lambda drv: f"{int(target_pct)} %" in drv.page_source
                         or f"{int(target_pct)}%" in drv.page_source
            )
            return True
        except Exception as e:
            self.log(f"✗ 等不到完成度 {target_pct}%（伺服器繁忙？逾時 {to}s）")
            save_debug_snapshot(d, f"progress_{target_pct}_timeout")
            return False

    def _click_done(self):
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        d = self.driver
        try:
            btn = WebDriverWait(d, 15).until(EC.element_to_be_clickable(
                (By.XPATH, "//*[(self::button or self::a or self::input) and (contains(.,'完成') or @value='完成')]")))
            btn.click()
            return True
        except Exception as e:
            self.log(f"✗ 完成失敗：{e}")
            return False

    # ── v1.0.22 完整模式輔助：依 profile 設定每個 radio group ──
    @staticmethod
    def _option_value(key, value_str):
        """選項字串 → ASP.NET radio value
           v1.0.32：hiva 對「沒有發生」/「從未做過」固定用 value=0，不是 index+1，需特別處理"""
        s = str(value_str or "").strip()
        # 特殊值優先（hiva ASP.NET 慣例）
        if s in ("沒有發生", "從未做過"):
            return "0"
        allowed = COMPLETE_FIELD_ALLOWED.get(key)
        if not allowed:
            return None
        if s in allowed:
            return str(allowed.index(s) + 1)
        if s == "是": return "1"
        if s == "否": return "2"
        if s == "不確定": return "3"
        return None

    def _get_questwizard_groups(self):
        """取得頁面上所有 QuestWizard radio 的 group name 列表（依 DOM 順序，去重）"""
        from selenium.webdriver.common.by import By
        radios = self.driver.find_elements(By.XPATH,
            "//input[@type='radio' and contains(@name,'QuestWizard')]")
        seen = set(); ordered = []
        for r in radios:
            n = r.get_attribute("name") or ""
            if n and n not in seen:
                seen.add(n); ordered.append(n)
        return ordered

    def _set_radio_for_group(self, name, value):
        """指定 group name + value 點選 radio"""
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            target = d.find_element(By.XPATH,
                f"//input[@type='radio' and @name=\"{name}\" and @value=\"{value}\"]")
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
            try: target.click()
            except Exception: d.execute_script("arguments[0].click();", target)
            return True
        except Exception:
            return False

    def _fill_groups_in_order(self, profile, field_keys, start_index=0):
        """依 DOM 順序給每個 group 設定（從 start_index 起對應 field_keys[0]）"""
        ordered = self._get_questwizard_groups()
        n_filled = 0
        for i, name in enumerate(ordered):
            if i < start_index:
                continue
            idx = i - start_index
            if idx >= len(field_keys):
                break
            key = field_keys[idx]
            val = profile.get(key)
            if val is None or val == "":
                # 缺值：依預設
                for k, _, dv, _ in COMPLETE_FIELDS:
                    if k == key:
                        val = dv; break
            radio_value = self._option_value(key, val)
            if radio_value is None:
                self.log(f"  ⚠ {key}={val} 沒有對應 value，跳過")
                continue
            if self._set_radio_for_group(name, radio_value):
                n_filled += 1
                self.dly.action()
            else:
                self.log(f"  ✗ {key}={val} (value={radio_value}) 點選失敗")
        return n_filled

    def _click_checkboxes_by_text(self, texts):
        """v1.0.23：勾選頁面上 label 文字含 texts 任一字的 checkbox（用於安非他命使用方式：吸入/注射/口服）"""
        if not texts: return 0
        from selenium.webdriver.common.by import By
        d = self.driver
        n = 0
        for t in texts:
            try:
                xp = f"//input[@type='checkbox' and (following-sibling::*[contains(.,'{t}')] or @value='{t}')]"
                els = d.find_elements(By.XPATH, xp)
                if not els:
                    xp2 = f"//*[contains(text(),'{t}')]/preceding-sibling::input[@type='checkbox'][1] | //*[contains(text(),'{t}')]/parent::*//input[@type='checkbox']"
                    els = d.find_elements(By.XPATH, xp2)
                if els:
                    el = els[0]
                    if not el.is_selected():
                        try: el.click()
                        except Exception: d.execute_script("arguments[0].click();", el)
                    n += 1
            except Exception:
                continue
        return n

    def _fill_text_input_by_name(self, name, value):
        """v1.0.32：以 name 屬性精準定位 input + 輸入文字"""
        if not value: return False
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            el = d.find_element(By.NAME, name)
            el.clear()
            el.send_keys(str(value))
            return True
        except Exception as e:
            self.log(f"  ⚠ 找不到 input name={name}：{e}")
            return False

    def _fill_text_input_by_label(self, label_text, value):
        """找 label 文字旁邊的 input[type=text]，輸入 value"""
        if not value: return False
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            xp = f"//*[contains(text(),'{label_text}')]/following::input[@type='text' or @type='email' or @type='tel'][1]"
            el = d.find_element(By.XPATH, xp)
            el.clear()
            el.send_keys(str(value))
            return True
        except Exception as e:
            self.log(f"  ⚠ 找不到 {label_text} 對應的 input：{e}")
            return False

    def _click_all_groups_no(self, expected_n=None, value="2"):
        """ASP.NET QuestWizard：頁面上每組 radio 由 name="ctl00$MainContent$QuestWizard$XXXX" 區隔；
           每組點 value=value（否=2 / 是=1）。
           expected_n：若指定，至少要點到這麼多組才算成功。"""
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            # 抓所有 QuestWizard radio，依 name 分組（保留 DOM 出現順序）
            radios = d.find_elements(By.XPATH,
                "//input[@type='radio' and contains(@name,'QuestWizard')]")
            seen = set()
            ordered_names = []
            for r in radios:
                n = r.get_attribute("name") or ""
                if n and n not in seen:
                    seen.add(n)
                    ordered_names.append(n)
            count = 0
            for name in ordered_names:
                # 該 group 的 value=否 那顆
                try:
                    target = d.find_element(By.XPATH,
                        f"//input[@type='radio' and @name=\"{name}\" and @value='{value}']")
                except Exception:
                    continue
                d.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
                try: target.click()
                except Exception: d.execute_script("arguments[0].click();", target)
                count += 1
                self.dly.action()
            if expected_n is not None and count < expected_n:
                self.log(f"⚠ 預期至少點 {expected_n} 組 radio，實際只點到 {count} 組")
                return False
            return count > 0
        except Exception as e:
            self.log(f"✗ click_all_groups_no 失敗：{e}")
            return False

    def _click_first_group_no(self, value="2"):
        """只點頁面上第一組 QuestWizard radio 的 value（給 Page 1/2/3 用）"""
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            radios = d.find_elements(By.XPATH,
                "//input[@type='radio' and contains(@name,'QuestWizard')]")
            if not radios:
                return False
            first_name = radios[0].get_attribute("name")
            target = d.find_element(By.XPATH,
                f"//input[@type='radio' and @name=\"{first_name}\" and @value='{value}']")
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
            try: target.click()
            except Exception: d.execute_script("arguments[0].click();", target)
            return True
        except Exception as e:
            self.log(f"✗ first group radio 點選失敗：{e}")
            return False

    def _dismiss_alert_if_any(self):
        """若頁面跳出 alert / Modal（例如「請於紅處填入您的答案」），按確定關閉"""
        from selenium.common.exceptions import NoAlertPresentException
        d = self.driver
        # JS alert
        try:
            a = d.switch_to.alert
            self.log(f"⚠ 偵測到瀏覽器 alert：{a.text[:80]}")
            a.accept()
            return True
        except NoAlertPresentException:
            pass
        except Exception:
            pass
        # 自訂 Modal（含 SweetAlert / Bootstrap modal）
        from selenium.webdriver.common.by import By
        for xp in [
            "//div[contains(@class,'swal') or contains(@class,'modal')]"
            "//button[contains(.,'確定') or contains(.,'OK') or contains(.,'關閉')]",
        ]:
            try:
                btn = d.find_element(By.XPATH, xp)
                if btn.is_displayed():
                    self.log("⚠ 偵測到自訂彈窗，按確定關閉")
                    btn.click()
                    return True
            except Exception:
                continue
        return False

    def _click_label_text(self, text):
        """Page 6/7：找文字含 text 的 label 點擊（含其前 radio）"""
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            els = d.find_elements(By.XPATH,
                f"//label[contains(.,'{text}')] | //*[contains(text(),'{text}')]/preceding-sibling::input[@type='radio'][1]")
            for el in els:
                try:
                    d.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    try: el.click()
                    except Exception: d.execute_script("arguments[0].click();", el)
                    return True
                except Exception:
                    continue
            return False
        except Exception as e:
            self.log(f"  label click 失敗 text={text}: {e}")
            return False

    def _select_by_id(self, elem_id, text):
        """用元素 ID 直接定位 <select>，依 visible text 選擇選項。
           Page 6 表單頭部有日曆 widget 的隱藏 select，會把 index 推掉，必須用 ID。"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import Select
        d = self.driver
        try:
            el = d.find_element(By.ID, elem_id)
            try:
                Select(el).select_by_visible_text(text)
                return True
            except Exception:
                for opt in el.find_elements(By.TAG_NAME, "option"):
                    if text == (opt.text or "").strip() or text in (opt.text or ""):
                        opt.click()
                        return True
                return False
        except Exception as e:
            self.log(f"✗ 找不到下拉 id={elem_id}：{e}")
            return False

    def _set_radio_by_name(self, name_attr, value):
        """以 name 屬性 + value 直接設定該題的 radio"""
        from selenium.webdriver.common.by import By
        d = self.driver
        try:
            xp = f"//input[@type='radio' and @name=\"{name_attr}\" and @value=\"{value}\"]"
            el = d.find_element(By.XPATH, xp)
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try: el.click()
            except Exception: d.execute_script("arguments[0].click();", el)
            return True
        except Exception as e:
            self.log(f"✗ radio name={name_attr} value={value} 點選失敗：{e}")
            return False

    def _wait_for_enabled(self, elem_id, timeout=None):
        """等到指定元素的 disabled 屬性消失（ASP.NET PostBack 完成的錨點）"""
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        d = self.driver
        to = timeout if timeout is not None else self.dly.wait_timeout
        try:
            WebDriverWait(d, to).until(
                lambda drv: not drv.find_element(By.ID, elem_id).get_attribute("disabled")
            )
            return True
        except Exception as e:
            self.log(f"✗ 等不到 {elem_id} 變 enabled：{e}")
            return False


# ── GUI ──────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        root.title(f"HIV 匿名諮詢代碼批次取號 v{VERSION}")
        root.geometry("1180x1050")
        root.minsize(960, 720)
        try:
            root.state("zoomed")
        except Exception:
            pass
        self.stop_evt  = threading.Event()
        self.pause_evt = threading.Event()
        self.results = []
        self.worker = None
        self.thread = None
        self._current_xlsx_path = None
        self._balancing = False
        self.output_dir_var = tk.StringVar(value=OUTPUT_DIR)
        self.theme_var = tk.StringVar(value=DEFAULT_THEME)
        self._spinner_running = False
        self._spinner_idx = 0
        self.mode_var = tk.StringVar(value="簡易（比例分布）")
        self._imported_profiles = []
        self._current_pool = []
        self._completed_index = 0
        self._build()
        self._apply_theme(DEFAULT_THEME)
        self._load_settings()
        # v1.0.21 啟動後檢查續傳
        self._check_resume()
        root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── v1.0.8 比例自動平衡：調一個其他自動降低，總和恆 100% ──
    def _balance_pct(self, block, changed_var):
        if self._balancing:
            return
        self._balancing = True
        try:
            try:
                new_val = int(changed_var.get())
            except Exception:
                new_val = 0
            new_val = max(0, min(100, new_val))
            if new_val != changed_var.get():
                changed_var.set(new_val)
            others = [(op, v, cnt) for (op, v, cnt) in block if v is not changed_var]
            if not others:
                return
            other_sum = sum(int(v.get() or 0) for _, v, _ in others)
            remain = 100 - new_val
            remain = max(0, remain)
            if other_sum > 0:
                # 比例縮放剩餘者；最後一個吃掉 rounding 差
                running = 0
                for (op, v, cnt) in others[:-1]:
                    nv = int(round(int(v.get() or 0) * remain / other_sum))
                    v.set(nv); running += nv
                others[-1][1].set(max(0, remain - running))
            else:
                # 其他全是 0：剩餘者平均分配
                n = len(others)
                base = remain // n
                rem  = remain - base * n
                for i, (op, v, cnt) in enumerate(others):
                    v.set(base + (1 if i < rem else 0))
        finally:
            self._balancing = False
        self._update_counts()

    def _balance_city(self, rows, changed_pv):
        """居住地動態列：和上面同樣邏輯，總和恆 100%"""
        if self._balancing:
            return
        self._balancing = True
        try:
            try:
                new_val = int(changed_pv.get())
            except Exception:
                new_val = 0
            new_val = max(0, min(100, new_val))
            if new_val != changed_pv.get():
                changed_pv.set(new_val)
            others = [(cv, pv, fr) for (cv, pv, fr) in rows if pv is not changed_pv]
            if not others:
                if new_val != 100:
                    changed_pv.set(100)
                return
            other_sum = sum(int(pv.get() or 0) for _, pv, _ in others)
            remain = max(0, 100 - new_val)
            if other_sum > 0:
                running = 0
                for (cv, pv, fr) in others[:-1]:
                    nv = int(round(int(pv.get() or 0) * remain / other_sum))
                    pv.set(nv); running += nv
                others[-1][1].set(max(0, remain - running))
            else:
                n = len(others)
                base = remain // n
                rem  = remain - base * n
                for i, (cv, pv, fr) in enumerate(others):
                    pv.set(base + (1 if i < rem else 0))
        finally:
            self._balancing = False
        self._update_counts()

    def _build(self):
        pad = {"padx": 6, "pady": 4}
        # ── v1.0.15 輸出路徑（最上方一列） ──
        path_fr = ttk.LabelFrame(self.root, text="輸出資料夾（log / Excel / debug 都會放這）")
        path_fr.pack(fill="x", **pad)
        ttk.Entry(path_fr, textvariable=self.output_dir_var, width=70).pack(side="left", padx=4, pady=4, fill="x", expand=True)
        ttk.Button(path_fr, text="📂 瀏覽…", command=self._pick_output_dir).pack(side="left", padx=2)
        ttk.Button(path_fr, text="✅ 套用", command=self._apply_output_dir).pack(side="left", padx=2)
        ttk.Button(path_fr, text="↩ 還原預設", command=self._reset_output_dir).pack(side="left", padx=2)
        ttk.Label(path_fr, text="  🎨 主題：").pack(side="left", padx=(12, 2))
        theme_cb = ttk.Combobox(path_fr, textvariable=self.theme_var, state="readonly",
                                 values=list(THEMES.keys()), width=14)
        theme_cb.pack(side="left", padx=2)
        theme_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_theme(self.theme_var.get()))

        # ── v1.0.19 Notebook：批次 / 單筆 兩個分頁 ──
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=6, pady=2)
        # v1.0.26：分頁切換時更新底部按鈕標籤
        self.notebook.bind("<<NotebookTabChanged>>", lambda e: self._update_start_btn_label())

        tab_batch  = ttk.Frame(self.notebook)
        tab_single = ttk.Frame(self.notebook)
        self.notebook.add(tab_batch,  text="📊 批次取號")
        self.notebook.add(tab_single, text="🎯 單筆取號")

        # ── v1.0.21 模式切換列（批次分頁最頂） ──
        mode_fr = ttk.LabelFrame(tab_batch, text="🎛 取號模式")
        mode_fr.pack(fill="x", padx=6, pady=4)
        ttk.Radiobutton(mode_fr, text="📊 簡易（比例分布隨機）", variable=self.mode_var,
                        value="簡易（比例分布）", command=self._on_mode_change).pack(side="left", padx=10, pady=6)
        ttk.Radiobutton(mode_fr, text="📋 完整（匯入 xlsx 逐筆指定）", variable=self.mode_var,
                        value="完整（匯入xlsx）", command=self._on_mode_change).pack(side="left", padx=10, pady=6)
        # v1.0.44：xlsx 工具列只在「完整模式」顯示，簡易模式不需要
        self._batch_xlsx_tools = ttk.Frame(mode_fr)
        ttk.Button(self._batch_xlsx_tools, text="📤 下載範例 xlsx",
                   command=self._export_sample_xlsx).pack(side="left", padx=4)
        ttk.Button(self._batch_xlsx_tools, text="📥 匯入 xlsx",
                   command=self._import_xlsx).pack(side="left", padx=4)
        self.import_status_var = tk.StringVar(value="（未匯入）")
        ttk.Label(self._batch_xlsx_tools, textvariable=self.import_status_var,
                  foreground="#1565c0", font=("微軟正黑體", 9, "bold")).pack(side="left", padx=12)
        # 簡易為 default → 啟動時不 pack（_on_mode_change 切換時才顯示）

        # ── 批次取號分頁內容 ──
        # v1.0.23：完整模式專用預覽列表（簡易模式時隱藏）
        preview_fr = ttk.LabelFrame(tab_batch, text="📋 已匯入預覽（顯示主要欄位）")
        cols = ("#", "性別", "出生年", "18歲前", "現居", "性傾向", "教育", "Q1性行為", "Q6性病", "Q7藥物", "P7篩檢")
        widths = (40, 60, 70, 90, 90, 70, 100, 70, 60, 60, 80)
        self.preview_tree = ttk.Treeview(preview_fr, columns=cols, show="headings", height=8)
        for c, w in zip(cols, widths):
            self.preview_tree.heading(c, text=c)
            self.preview_tree.column(c, width=w, anchor="center")
        self.preview_tree.pack(fill="x", padx=4, pady=4)
        self._batch_preview_fr = preview_fr
        # 簡易模式時不要 pack 顯示

        topbar = ttk.Frame(tab_batch)
        topbar.pack(fill="x", **pad)
        self._batch_topbar = topbar  # 完整模式時要隱藏
        midbar = ttk.Frame(tab_batch)
        midbar.pack(fill="x", **pad)
        self._batch_midbar = midbar  # 完整模式時要隱藏
        col_left  = ttk.Frame(midbar)
        col_right = ttk.Frame(midbar)
        col_left.pack(side="left",  fill="both", expand=True, padx=(0, 4))
        col_right.pack(side="left", fill="both", expand=True, padx=(4, 0))

        # ─ 基本設定 ─
        top = ttk.LabelFrame(topbar, text="基本設定（批次）")
        top.pack(side="left", padx=(0, 4))
        ttk.Label(top, text="總筆數：").grid(row=0, column=0, padx=4, pady=4, sticky="w")
        self.total_var = tk.IntVar(value=30)
        e_total = self._mk_int_entry(top, self.total_var, width=8, justify="center")
        e_total.grid(row=0, column=1, padx=4, pady=4, sticky="w")
        self.total_var.trace_add("write", lambda *a: self._update_counts())
        ttk.Label(top, text="出生年：").grid(row=0, column=2, padx=12, pady=4, sticky="w")
        self.year_lo = tk.IntVar(value=1971)
        self.year_hi = tk.IntVar(value=2008)
        self._mk_int_entry(top, self.year_lo, width=6, justify="center").grid(row=0, column=3, padx=2, pady=4)
        ttk.Label(top, text=" ~ ").grid(row=0, column=4)
        self._mk_int_entry(top, self.year_hi, width=6, justify="center").grid(row=0, column=5, padx=2, pady=4)

        # ─ 速度 / 延遲設定（v1.0.50：Claude Design 概念稿 3.2 節奏控制）─
        speed = ttk.LabelFrame(topbar, text="節奏控制（按一個 preset 即可，進階自訂可展開）")
        speed.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # 4 顆 preset chip 按鈕
        self.speed_preset = tk.StringVar(value="中（推薦）")
        chips = [("🐇 快", "快（偷懶模式）"),
                 ("🚶 中", "中（推薦）"),
                 ("🐢 慢", "慢（高峰時段）"),
                 ("⏳ 極慢", "極慢（嚴重塞車）")]
        chip_fr = ttk.Frame(speed)
        chip_fr.grid(row=0, column=0, columnspan=5, padx=4, pady=(8, 4), sticky="w")
        ttk.Label(chip_fr, text="速度：", font=("微軟正黑體", 10, "bold")).pack(side="left", padx=(0, 8))
        self._chip_btns = {}
        def _click_chip(value):
            self.speed_preset.set(value)
            self._apply_speed_preset(None)
            _update_chip_state()
        for label, value in chips:
            btn = ttk.Button(chip_fr, text=label, width=8,
                              command=lambda v=value: _click_chip(v))
            btn.pack(side="left", padx=2)
            self._chip_btns[value] = btn
        # active chip 視覺：用 state="pressed" 模擬，或維持普通；簡單用 text 加標
        def _update_chip_state():
            cur = self.speed_preset.get()
            for v, btn in self._chip_btns.items():
                if v == cur:
                    btn.state(["pressed"])
                else:
                    btn.state(["!pressed"])
        _update_chip_state()
        # 顯示目前預設名稱（自訂時會出現「自訂」）
        self._speed_label = ttk.Label(chip_fr, textvariable=self.speed_preset,
                                       foreground="#7d8696", font=("Consolas", 9))
        self._speed_label.pack(side="left", padx=(12, 0))

        # 進階區間設定（disclosure：點開才顯示 6 個 entry + 等待逾時）
        # v1.0.57：等待逾時也移進 adv，預設只看到 chip
        self._adv_visible = False
        self._adv_btn_text = tk.StringVar(value="▶ 進階設定（自訂秒數 / 逾時）")
        adv_btn = ttk.Button(speed, textvariable=self._adv_btn_text,
                              command=lambda: _toggle_adv())
        adv_btn.grid(row=1, column=0, columnspan=5, padx=4, pady=(6, 4), sticky="w")

        adv_fr = ttk.Frame(speed)
        self.dly_act_lo = tk.DoubleVar(value=0.3)
        self.dly_act_hi = tk.DoubleVar(value=0.7)
        self.dly_page_lo = tk.DoubleVar(value=0.8)
        self.dly_page_hi = tk.DoubleVar(value=1.5)
        self.dly_btw_lo = tk.DoubleVar(value=1.5)
        self.dly_btw_hi = tk.DoubleVar(value=3.0)
        self.dly_timeout = tk.IntVar(value=30)
        # 動作延遲
        ttk.Label(adv_fr, text="動作延遲：").grid(row=0, column=0, padx=4, pady=2, sticky="w")
        self._mk_float_entry(adv_fr, self.dly_act_lo, width=6).grid(row=0, column=1, padx=2)
        ttk.Label(adv_fr, text="~").grid(row=0, column=2)
        self._mk_float_entry(adv_fr, self.dly_act_hi, width=6).grid(row=0, column=3, padx=2)
        ttk.Label(adv_fr, text="（每點 radio/select 後）", foreground="#666").grid(row=0, column=4, padx=8, sticky="w")
        # 頁面切換延遲
        ttk.Label(adv_fr, text="頁面切換延遲：").grid(row=1, column=0, padx=4, pady=2, sticky="w")
        self._mk_float_entry(adv_fr, self.dly_page_lo, width=6).grid(row=1, column=1, padx=2)
        ttk.Label(adv_fr, text="~").grid(row=1, column=2)
        self._mk_float_entry(adv_fr, self.dly_page_hi, width=6).grid(row=1, column=3, padx=2)
        ttk.Label(adv_fr, text="（按下一步後）", foreground="#666").grid(row=1, column=4, padx=8, sticky="w")
        # 每筆間隔
        ttk.Label(adv_fr, text="每筆間隔：").grid(row=2, column=0, padx=4, pady=2, sticky="w")
        self._mk_float_entry(adv_fr, self.dly_btw_lo, width=6).grid(row=2, column=1, padx=2)
        ttk.Label(adv_fr, text="~").grid(row=2, column=2)
        self._mk_float_entry(adv_fr, self.dly_btw_hi, width=6).grid(row=2, column=3, padx=2)
        ttk.Label(adv_fr, text="（取完一筆到下一筆開始）", foreground="#666").grid(row=2, column=4, padx=8, sticky="w")
        # 等待逾時 v1.0.57 移進來
        ttk.Label(adv_fr, text="等待逾時：").grid(row=3, column=0, padx=4, pady=2, sticky="w")
        self._mk_int_entry(adv_fr, self.dly_timeout, width=6).grid(row=3, column=1, padx=2)
        ttk.Label(adv_fr, text="秒  （多人連線會慢，建議 30+）",
                  foreground="#666").grid(row=3, column=2, columnspan=3, padx=8, sticky="w")

        def _toggle_adv():
            if self._adv_visible:
                adv_fr.grid_forget()
                self._adv_btn_text.set("▶ 進階設定（自訂秒數 / 逾時）")
                self._adv_visible = False
            else:
                adv_fr.grid(row=2, column=0, columnspan=5, padx=4, pady=(0, 4), sticky="ew")
                self._adv_btn_text.set("▼ 進階設定（自訂秒數 / 逾時）")
                self._adv_visible = True

        # 監聽 self.speed_preset 變化，更新 chip active 視覺（_apply_speed_preset 會被外部 set）
        self.speed_preset.trace_add("write", lambda *a: _update_chip_state())

        # ─ 比例設定區（2 欄並列） ─
        # 左欄：性別 / 國籍 / 教育
        # v1.0.49：每維度自己的色階（Claude Design exe_concept.html 3.3）
        self.gender_pcts  = self._make_pct_block("性別分布", GENDERS, [100, 0, 0, 0],
                                                  palette=["#5b8def", "#a4c0f0", "#dde5f4", "#eef1f5"],
                                                  parent=col_left)
        self.nation_pcts  = self._make_pct_block("國籍分布", NATIONS, [100, 0],
                                                  palette=["#5fa787", "#9bc8b3"],
                                                  parent=col_left)
        self.edu_pcts     = self._make_pct_block("教育程度分布", EDUS, [0, 33, 34, 33, 0],
                                                  palette=["#6c5ce7", "#a092ee", "#cbc4f4", "#e3dff8", "#eef1f5"],
                                                  parent=col_left)
        # 右欄：性傾向 / 篩檢習慣 / 18 歲前居住地 / 現居住地
        self.orient_pcts  = self._make_pct_block("性傾向分布", ORIENTS, [0, 0, 100],
                                                  palette=["#d97a8a", "#e9a5b1", "#f3cbd2"],
                                                  parent=col_right)
        self.testing_pcts = self._make_pct_block("篩檢習慣分布（Q1）", TESTING, [50, 50],
                                                  palette=["#c39145", "#dcb888"],
                                                  parent=col_right)
        self.res18_rows   = self._make_city_block("18 歲以前居住地分布",                          parent=col_right)
        self.resCur_rows  = self._make_city_block("現居住地分布",                                  parent=col_right)

        # ── v1.0.19 單筆取號分頁內容 ──
        self._build_single_tab(tab_single)

        # ─ 控制按鈕 ─
        btn_fr = ttk.Frame(self.root)
        btn_fr.pack(fill="x", padx=6, pady=8)

        # 左側：輔助按鈕 + 路徑說明
        left = ttk.Frame(btn_fr)
        left.pack(side="left", fill="y")
        # v1.0.44：簡化為單一「輸出資料夾」放大按鈕；log/debug 子資料夾隱藏不暴露給使用者
        ttk.Button(left, text="📁 輸出資料夾", command=self.open_outdir,
                   style="Medium.TButton").pack(side="left", padx=2)
        ttk.Label(left, text=f"  輸出：{OUTPUT_DIR}", foreground="#666").pack(side="left", padx=8)

        # 右側：大按鈕（開始 / 暫停 / 停止）
        # v1.0.57：改用 tk.Button 確保三顆視覺尺寸一致（ttk 在某些 Windows 主題下吃不到 padding/bg）
        right = ttk.Frame(btn_fr)
        right.pack(side="right")
        big_style = ttk.Style()
        try:
            # 保留 Medium.TButton 給「輸出資料夾」用
            big_style.configure("Medium.TButton", font=("微軟正黑體", 11, "bold"), padding=(14, 8))
        except Exception:
            pass
        # 三顆同尺寸（width=10 字寬、padx/pady 一致）；只用色彩區分
        _BIG_FONT = ("微軟正黑體", 14, "bold")
        _BTN_W = 10
        self.start_btn = tk.Button(right, text="▶ 開始取號", command=self._smart_start,
                                    font=_BIG_FONT, width=_BTN_W,
                                    bg="#2c6bd1", fg="#ffffff",
                                    activebackground="#1f4f9c", activeforeground="#ffffff",
                                    disabledforeground="#ffffff",
                                    relief="flat", bd=0, padx=10, pady=10, cursor="hand2")
        self.start_btn.pack(side="left", padx=6)
        self.pause_btn = tk.Button(right, text="⏸ 暫停", command=self.toggle_pause,
                                    state="disabled",
                                    font=_BIG_FONT, width=_BTN_W,
                                    bg="#fff3e0", fg="#e65100",
                                    activebackground="#ffe0b2", activeforeground="#bf360c",
                                    disabledforeground="#bdbdbd",
                                    relief="flat", bd=0, padx=10, pady=10, cursor="hand2")
        self.pause_btn.pack(side="left", padx=6)
        self.stop_btn = tk.Button(right, text="■ 停止並關閉", command=self.stop,
                                   state="disabled",
                                   font=_BIG_FONT, width=_BTN_W,
                                   bg="#fbe9e7", fg="#b71c1c",
                                   activebackground="#ffcdd2", activeforeground="#7f0000",
                                   disabledforeground="#bdbdbd",
                                   relief="flat", bd=0, padx=10, pady=10, cursor="hand2")
        self.stop_btn.pack(side="left", padx=6)

        # ─ 進度條（v1.0.19 加高 + % 在條內） ─
        prog_fr = ttk.Frame(self.root)
        prog_fr.pack(fill="x", padx=6, pady=2)
        self.spinner_label = ttk.Label(prog_fr, text="●", font=("微軟正黑體", 13, "bold"), foreground="#90a4ae")
        self.spinner_label.pack(side="left", padx=(0, 6))
        self.progress_var = tk.StringVar(value="待命")
        ttk.Label(prog_fr, textvariable=self.progress_var,
                  font=("微軟正黑體", 11, "bold")).pack(side="left")
        # v1.0.19/v1.0.45/v1.0.48：自訂 Canvas 進度條（高度 58，Canvas 自繪卡通跑者 + 終點旗）
        _theme_now = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
        self.pb = CanvasProgressBar(prog_fr, height=58, bg=_theme_now["panel"],
                                     accent=_theme_now["accent"], accent_2=_theme_now["accent_hover"])
        self.pb.pack(side="left", padx=12, fill="x", expand=True)
        # v1.0.19 統計列（置中）
        stat_fr = ttk.Frame(self.root)
        stat_fr.pack(fill="x", padx=6, pady=(0, 4))
        # 左側：自動開啟 Excel checkbox
        self.auto_open_xlsx = tk.BooleanVar(value=True)
        ttk.Checkbutton(stat_fr, text="完成後自動停止並開啟 Excel",
                        variable=self.auto_open_xlsx).pack(side="right", padx=8)
        # 中央：統計
        self.stats_var = tk.StringVar(value="平均 — 秒/筆   |   已用 00m00s   |   剩餘 —   |   預計完成 —")
        self.stats_label = ttk.Label(stat_fr, textvariable=self.stats_var,
                  font=("微軟正黑體", 11, "bold"), foreground="#1565c0", anchor="center")
        self.stats_label.pack(side="left", fill="x", expand=True)

        log_fr = ttk.LabelFrame(self.root, text="執行紀錄（綠=成功 / 黃=注意 / 紅=錯誤）")
        log_fr.pack(fill="both", expand=True, padx=6, pady=6)
        self.log_box = tk.Text(log_fr, height=15, font=("Consolas", 10), bg="#1e1e1e", fg="#dcdcdc",
                               insertbackground="#dcdcdc")
        self.log_box.pack(fill="both", expand=True, padx=4, pady=(4, 0))
        # log 著色 tag
        self.log_box.tag_configure("success", foreground="#4caf50")
        self.log_box.tag_configure("warn",    foreground="#ffb300")
        self.log_box.tag_configure("error",   foreground="#ef5350", font=("Consolas", 10, "bold"))
        self.log_box.tag_configure("info",    foreground="#dcdcdc")
        # log 控制按鈕
        log_btns = ttk.Frame(log_fr)
        log_btns.pack(fill="x", padx=4, pady=4)
        ttk.Button(log_btns, text="🧹 清除 Log",  command=self.clear_log).pack(side="left", padx=2)
        ttk.Button(log_btns, text="💾 儲存 Log",  command=self.save_log).pack(side="left", padx=2)
        # v1.0.57：全展開（Toplevel 全螢幕看完整 log）+ 一鍵複製到剪貼簿
        ttk.Button(log_btns, text="🔍 全展開",   command=self.show_log_full).pack(side="left", padx=2)
        ttk.Button(log_btns, text="📋 複製全部", command=self.copy_log_clipboard).pack(side="left", padx=2)
        ttk.Label(log_btns, text="（儲存 Log 會將目前畫面內容存成 .txt，方便回報修正）",
                  foreground="#666").pack(side="left", padx=12)

        # ── v1.0.45 免責聲明 / v1.0.51 兩欄：左版本、右免責聲明（拿掉 build 日期）──
        disclaimer_fr = ttk.Frame(self.root)
        disclaimer_fr.pack(fill="x", padx=14, pady=(0, 6))
        ttk.Label(disclaimer_fr,
                  text=f"v{VERSION}",
                  foreground="#9e9e9e",
                  font=("Consolas", 9)).pack(side="left")
        ttk.Label(disclaimer_fr,
                  text="⚠ 免責聲明：本工具僅供研究網頁服務技術用途",
                  foreground="#9e9e9e",
                  font=("微軟正黑體", 9)).pack(side="right")

    def _make_pct_block(self, title, options, defaults, parent=None, palette=None):
        """v1.0.49 dashboard 卡片版（取代 v1.0.x LabelFrame）：
        頂部 title + 「∑ N%」狀態 → 水平堆疊比例條 → legend 列（色點 + 標籤 + entry + 筆數）
        參考 Claude Design 概念稿 exe_concept.html 3.3 人物輪廓卡。
        palette 是該維度專屬色階（從主到淡），不夠長會重複用最後一格。"""
        if not palette:
            palette = ["#5b8def", "#a4c0f0", "#dde5f4", "#e3e8ef", "#eef1f5"]
        # 不夠長補淡灰
        while len(palette) < len(options):
            palette = palette + ["#e3e8ef"]

        # 卡容器
        card = ttk.Frame(parent or self.root, padding=(12, 8))
        card.pack(fill="x", padx=6, pady=4)

        # Header：標題 + 總和狀態（滿 100% 時灰，否則橘色提示）
        header = ttk.Frame(card)
        header.pack(fill="x")
        ttk.Label(header, text=title, font=("微軟正黑體", 11, "bold")).pack(side="left")
        sum_var = tk.StringVar(value="")
        sum_lbl = ttk.Label(header, textvariable=sum_var, font=("Consolas", 9))
        sum_lbl.pack(side="right")

        # 水平堆疊比例條
        bar = tk.Canvas(card, height=8, bg="#eef1f5", highlightthickness=0, bd=0)
        bar.pack(fill="x", pady=(6, 6))

        # Legend 列
        rows_fr = ttk.Frame(card)
        rows_fr.pack(fill="x")
        vars_ = []
        bar_segs = []
        for i, (op, dv) in enumerate(zip(options, defaults)):
            c = palette[i]
            row = ttk.Frame(rows_fr)
            row.pack(fill="x", pady=1)
            # 色點（用 ●，前景色 = palette）
            ttk.Label(row, text="●", foreground=c, font=("微軟正黑體", 10)).pack(side="left", padx=(0, 4))
            ttk.Label(row, text=op, width=12, anchor="w").pack(side="left")
            v = tk.IntVar(value=dv)
            self._mk_int_entry(row, v, width=5, justify="center").pack(side="left")
            ttk.Label(row, text="%").pack(side="left", padx=(0, 8))
            cnt = tk.StringVar(value="0 筆")
            ttk.Label(row, textvariable=cnt, foreground="#7d8696",
                      font=("Consolas", 9), width=8, anchor="w").pack(side="left")
            vars_.append((op, v, cnt))
            bar_segs.append((c, v))

        # 重畫堆疊比例條 + 更新「∑」狀態
        def redraw_bar(*_a):
            try:
                bar.delete("all")
                w = bar.winfo_width()
                if w <= 1:  # 還沒 layout 完
                    return
                x = 0
                total_pct = sum(v.get() for _, v, _ in vars_)
                for c, v in bar_segs:
                    if total_pct <= 0:
                        break
                    seg_w = w * max(0, v.get()) / 100.0
                    if seg_w > 0:
                        bar.create_rectangle(x, 0, x + seg_w, 8, fill=c, outline="")
                        x += seg_w
                if total_pct == 100:
                    sum_var.set("")
                else:
                    sum_var.set(f"⚠ ∑ {total_pct}%")
            except Exception:
                pass

        bar.bind("<Configure>", redraw_bar)
        for (op, v, cnt) in vars_:
            # rebalance + redraw
            v.trace_add("write", lambda *a, _v=v, _block=vars_:
                        (self._balance_pct(_block, _v), redraw_bar()))
        # 初次呼叫一次（在 layout 完才有效，所以用 after）
        card.after(50, redraw_bar)
        return vars_

    # ── v1.0.19/25 單筆取號分頁 ──
    def _build_single_tab(self, parent):
        # v1.0.25：模式切換（單筆也分簡易/完整）
        self.single_mode_var = tk.StringVar(value="簡易")
        mode_fr = ttk.LabelFrame(parent, text="🎛 單筆取號模式")
        mode_fr.pack(fill="x", padx=6, pady=(8, 4))
        ttk.Radiobutton(mode_fr, text="📊 簡易（8 個基本欄位）", variable=self.single_mode_var,
                        value="簡易", command=self._on_single_mode_change).pack(side="left", padx=10, pady=6)
        ttk.Radiobutton(mode_fr, text="📋 完整（全部 48 欄精準設定）", variable=self.single_mode_var,
                        value="完整", command=self._on_single_mode_change).pack(side="left", padx=10, pady=6)
        # v1.0.26：移除上方獨立按鈕；單筆 / 批次 都用底部「▶ 開始取號」（會依當前分頁切換）
        self.single_result_var = tk.StringVar(value="尚未取號")
        ttk.Label(mode_fr, textvariable=self.single_result_var,
                  font=("微軟正黑體", 12, "bold"), foreground="#1565c0").pack(side="right", padx=10)
        ttk.Label(mode_fr, text="（按下方「▶ 開始取號」啟動）  ",
                  foreground="#666", font=("微軟正黑體", 9, "italic")).pack(side="right")

        self.single_vars = {}

        # v1.0.25：同 key 重用 var → 簡易/完整雙向同步
        def get_or_create_var(key, default, var_class=tk.StringVar):
            if key in self.single_vars:
                return self.single_vars[key]
            v = var_class(value=default)
            self.single_vars[key] = v
            return v

        def mk_radio_row(parent, label, key, options, default):
            fr = ttk.Frame(parent); fr.pack(fill="x", pady=2, padx=6)
            ttk.Label(fr, text=label, width=18, anchor="w").pack(side="left")
            v = get_or_create_var(key, default)
            for op in options:
                ttk.Radiobutton(fr, text=op, variable=v, value=op).pack(side="left", padx=3)

        def mk_combo_row(parent, label, key, values, default, width=22):
            fr = ttk.Frame(parent); fr.pack(fill="x", pady=2, padx=6)
            ttk.Label(fr, text=label, width=18, anchor="w").pack(side="left")
            v = get_or_create_var(key, default)
            ttk.Combobox(fr, textvariable=v, values=values, state="readonly", width=width).pack(side="left", padx=4)

        def mk_year_row(parent, label, key, default):
            fr = ttk.Frame(parent); fr.pack(fill="x", pady=2, padx=6)
            ttk.Label(fr, text=label, width=18, anchor="w").pack(side="left")
            v = get_or_create_var(key, default, var_class=tk.IntVar)
            self._mk_int_entry(fr, v, width=8).pack(side="left", padx=4)
            ttk.Label(fr, text=" 西元年").pack(side="left", padx=4)

        def mk_text_row(parent, label, key, default=""):
            fr = ttk.Frame(parent); fr.pack(fill="x", pady=2, padx=6)
            ttk.Label(fr, text=label, width=18, anchor="w").pack(side="left")
            v = get_or_create_var(key, default)
            ttk.Entry(fr, textvariable=v, width=30).pack(side="left", padx=4)

        # ── 簡易模式 frame ──
        self.single_simple_fr = ttk.Frame(parent)
        s_left  = ttk.LabelFrame(self.single_simple_fr, text="個案資料")
        s_right = ttk.LabelFrame(self.single_simple_fr, text="說明")
        s_left.pack(side="left",  fill="both", expand=True, padx=(6, 3))
        s_right.pack(side="left", fill="both", expand=True, padx=(3, 6))
        mk_radio_row(s_left, "性別",       "gender", GENDERS, "男")
        mk_radio_row(s_left, "國籍",       "nation", NATIONS, "本國籍")
        mk_year_row (s_left, "出生年",     "year",   1990)
        mk_combo_row(s_left, "18歲前居住地","res18",  CITIES, "台南市")
        mk_combo_row(s_left, "現居住地",   "resCur", CITIES, "台南市")
        mk_radio_row(s_left, "性傾向",     "orient", ORIENTS, "異性")
        mk_radio_row(s_left, "教育程度",   "edu",    EDUS,    "高中職")
        mk_radio_row(s_left, "篩檢習慣",   "testing",TESTING, "否")
        ttk.Label(s_right, text="""
簡易模式：
  ‧ 8 個基本欄位
  ‧ Q1-Q13 全部用「否」走最簡單路徑
  ‧ 適合大部分外展協助情境

完整模式：
  ‧ 切到右上「完整」可精準設定全部 48 欄
  ‧ 適合需要 Q6=是 / Q7=是 / PrEP-PEP 等特殊條件
""", justify="left", padding=(10, 10), foreground="#37474f").pack(anchor="w", padx=8, pady=8)

        # ── 完整模式 frame（分頁顯示） ──
        self.single_complete_fr = ttk.Frame(parent)
        # 用內部 Notebook 分 7 頁
        sub_nb = ttk.Notebook(self.single_complete_fr)
        sub_nb.pack(fill="both", expand=True, padx=6, pady=4)

        # P1
        p1 = ttk.Frame(sub_nb); sub_nb.add(p1, text="P1 風險評估")
        mk_radio_row(p1, "Q1 過性行為",      "q1_sex",      ["是", "否"], "否")
        mk_radio_row(p1, "Q2 全程保險套",    "q2_condom",   ["是", "否", "沒有發生"], "沒有發生")
        mk_radio_row(p1, "Q3 跟固定性伴侶",  "q3_regular",  ["是", "否", "沒有發生"], "沒有發生")
        mk_radio_row(p1, "Q4 用酒",          "q4_alcohol",  ["是", "否", "沒有發生"], "沒有發生")
        mk_radio_row(p1, "Q5 用藥",          "q5_drug",     ["是", "否", "沒有發生"], "沒有發生")

        # P2
        p2 = ttk.Frame(sub_nb); sub_nb.add(p2, text="P2 性病")
        mk_radio_row(p2, "Q6 1年內感染性病",     "q6_std",       ["是", "否"], "否")
        mk_radio_row(p2, "Q6.1 HIV (Q6=是才填)", "q6_hiv",       ["是", "否"], "否")
        mk_text_row (p2, "  HIV 想再篩原因",     "q6_hiv_reason")
        mk_radio_row(p2, "Q6.2 菜花",            "q6_warts",     ["是", "否"], "否")
        mk_radio_row(p2, "Q6.3 梅毒",            "q6_syphilis",  ["是", "否"], "否")
        mk_radio_row(p2, "Q6.4 淋病",            "q6_gonorrhea", ["是", "否"], "否")
        mk_radio_row(p2, "Q6.5 披衣菌",          "q6_chlamydia", ["是", "否"], "否")
        mk_radio_row(p2, "Q6.6 疱疹",            "q6_herpes",    ["是", "否"], "否")
        mk_radio_row(p2, "Q6.7 A肝",             "q6_hepA",      ["是", "否"], "否")
        mk_radio_row(p2, "Q6.8 C肝",             "q6_hepC",      ["是", "否"], "否")
        mk_text_row (p2, "Q6.9 其他",            "q6_other")

        # P3
        p3 = ttk.Frame(sub_nb); sub_nb.add(p3, text="P3 藥物")
        mk_radio_row(p3, "Q7 1年內成癮藥物",     "q7_drug_use",  ["是", "否"], "否")
        mk_radio_row(p3, "Q7.1 安非他命 (Q7=是)", "q7_amph",     ["是", "否"], "否")
        mk_text_row (p3, "  使用方式(吸入,注射,口服)", "q7_amph_method")
        mk_radio_row(p3, "Q7.2 G水",             "q7_ghb",       ["是", "否"], "否")
        mk_radio_row(p3, "Q7.3 搖頭丸",          "q7_mdma",      ["是", "否"], "否")
        mk_radio_row(p3, "Q7.4 K他命",           "q7_ketamine",  ["是", "否"], "否")
        mk_radio_row(p3, "Q7.5 RUSH",            "q7_rush",      ["是", "否"], "否")
        mk_radio_row(p3, "Q7.6 喵喵",            "q7_meph",      ["是", "否"], "否")
        mk_radio_row(p3, "Q7.7 海洛因",          "q7_heroin",    ["是", "否"], "否")
        mk_radio_row(p3, "Q7.8 大麻",            "q7_marijuana", ["是", "否"], "否")
        mk_text_row (p3, "Q7.9 其他",            "q7_other")
        mk_radio_row(p3, "Q7.10 目前使用狀態",   "q7_status",    ["還在使用", "已停用"], "")

        # P4
        p4 = ttk.Frame(sub_nb); sub_nb.add(p4, text="P4 性接觸場合")
        mk_radio_row(p4, "Q8a 網路認識",         "q8a_online",      ["是", "否"], "否")
        mk_radio_row(p4, "Q8b 娛樂場所認識",     "q8b_venue",       ["是", "否"], "否")
        mk_radio_row(p4, "Q8c 性交易服務者",     "q8c_sex_worker",  ["是", "否"], "否")
        mk_radio_row(p4, "Q8d 性交易消費者",     "q8d_sex_consumer",["是", "否"], "否")

        # P5
        p5 = ttk.Frame(sub_nb); sub_nb.add(p5, text="P5 PEP/PrEP")
        mk_radio_row(p5, "Q9 固定伴侶HIV",       "q9_partner_hiv",
                      ["是", "否", "不確定", "目前沒有固定性伴侶"], "否")
        mk_radio_row(p5, "Q10 PEP使用過",        "q10_pep_used",   ["是", "否"], "否")
        mk_radio_row(p5, "Q11 PEP想服用",        "q11_pep_want",   ["是", "否"], "否")
        mk_radio_row(p5, "Q12 聽過PrEP",         "q12_prep_heard", ["是", "否"], "否")
        mk_radio_row(p5, "Q13 PrEP想服用",       "q13_prep_want",  ["是", "否"], "否")

        # P6（基本資料）— v1.0.25：直接使用編輯控件，與簡易模式雙向同步（共用同一 StringVar）
        p6 = ttk.Frame(sub_nb); sub_nb.add(p6, text="P6 基本資料")
        ttk.Label(p6, text="✓ 與簡易模式雙向同步 — 在此修改會同步反映到簡易分頁",
                  foreground="#2e7d32", font=("微軟正黑體", 9, "italic")).pack(pady=(6, 4))
        mk_radio_row(p6, "性別",        "gender", GENDERS, "男")
        mk_radio_row(p6, "國籍",        "nation", NATIONS, "本國籍")
        mk_year_row (p6, "出生年",      "year",   1990)
        mk_combo_row(p6, "18歲前居住地","res18",  CITIES, "台南市")
        mk_combo_row(p6, "現居住地",    "resCur", CITIES, "台南市")
        mk_radio_row(p6, "性傾向",      "orient", ORIENTS, "異性")
        mk_radio_row(p6, "教育程度",    "edu",    EDUS,    "高中職")

        # P7
        p7 = ttk.Frame(sub_nb); sub_nb.add(p7, text="P7 篩檢習慣 + 聯絡")
        mk_radio_row(p7, "Q1 篩檢習慣", "testing_habit", ["是", "否", "從未做過"], "否")
        mk_text_row (p7, "手機號碼（選填）", "phone")
        mk_text_row (p7, "E-mail（選填）",   "email")
        mk_text_row (p7, "其它聯絡（選填）", "other_contact")

        # 預設顯示簡易
        self._on_single_mode_change()

    def _on_single_mode_change(self):
        m = self.single_mode_var.get()
        if m == "完整":
            try: self.single_simple_fr.pack_forget()
            except Exception: pass
            self.single_complete_fr.pack(fill="both", expand=True)
        else:
            try: self.single_complete_fr.pack_forget()
            except Exception: pass
            self.single_simple_fr.pack(fill="both", expand=True)

    def _make_city_block(self, title, parent=None):
        """v1.0.51 dashboard 卡片版（與 _make_pct_block 同節奏）：
        Header（標題 + 添加按鈕）→ 水平堆疊比例條 → 動態列（●+Combobox+%+✕）
        城市色階：青色系（Claude Design 概念稿 居住地 #0fa3a3 / #5cc4c4 / #a8dede）
        rows 結構保留 [(StringVar, IntVar, Frame), ...]，外部 _balance_city 不用改。"""
        palette = ["#0fa3a3", "#5cc4c4", "#a8dede", "#cfeaea", "#e8f4f4"]

        card = ttk.Frame(parent or self.root, padding=(12, 8))
        card.pack(fill="x", padx=6, pady=4)

        # Header
        header = ttk.Frame(card)
        header.pack(fill="x")
        ttk.Label(header, text=title, font=("微軟正黑體", 11, "bold")).pack(side="left")
        sum_var = tk.StringVar(value="")
        ttk.Label(header, textvariable=sum_var, font=("Consolas", 9)).pack(side="right", padx=(0, 6))

        # 水平堆疊比例條
        bar = tk.Canvas(card, height=8, bg="#eef1f5", highlightthickness=0, bd=0)
        bar.pack(fill="x", pady=(6, 6))

        rows_holder = ttk.Frame(card)
        rows_holder.pack(fill="x")
        rows = []

        def redraw_bar(*_a):
            try:
                bar.delete("all")
                w = bar.winfo_width()
                if w <= 1:
                    return
                x = 0
                total_pct = 0
                for i, (cv, pv, _) in enumerate(rows):
                    try: total_pct += pv.get()
                    except Exception: pass
                for i, (cv, pv, _) in enumerate(rows):
                    try: p = max(0, pv.get())
                    except Exception: p = 0
                    seg_w = w * p / 100.0
                    if seg_w > 0:
                        c = palette[i % len(palette)]
                        bar.create_rectangle(x, 0, x + seg_w, 8, fill=c, outline="")
                        x += seg_w
                if total_pct == 100:
                    sum_var.set("")
                else:
                    sum_var.set(f"⚠ ∑ {total_pct}%")
            except Exception:
                pass

        bar.bind("<Configure>", redraw_bar)

        def add_row(city="台南市", pct=100):
            row_fr = ttk.Frame(rows_holder)
            row_fr.pack(fill="x", pady=1)
            cv = tk.StringVar(value=city)
            pv = tk.IntVar(value=pct)
            # 色點（依目前 row 索引取 palette 色）
            idx = len(rows)
            c = palette[idx % len(palette)]
            ttk.Label(row_fr, text="●", foreground=c, font=("微軟正黑體", 10)).pack(side="left", padx=(0, 4))
            cb = ttk.Combobox(row_fr, textvariable=cv, values=CITIES, width=12, state="readonly")
            cb.pack(side="left", padx=2)
            self._mk_int_entry(row_fr, pv, width=5, justify="center").pack(side="left", padx=2)
            ttk.Label(row_fr, text="%").pack(side="left", padx=(0, 8))
            def remove():
                row_fr.destroy()
                rows.remove(item)
                if rows and not self._balancing:
                    self._balance_city(rows, rows[0][1])
                self._update_counts()
                redraw_bar()
            ttk.Button(row_fr, text="✕", width=3, command=remove).pack(side="left", padx=4)
            item = (cv, pv, row_fr)
            rows.append(item)
            pv.trace_add("write", lambda *a, _pv=pv, _rows=rows:
                         (self._balance_city(_rows, _pv), redraw_bar()))
            cv.trace_add("write", lambda *a: redraw_bar())
            redraw_bar()

        # 添加按鈕
        ttk.Button(header, text="＋ 添加縣市", command=lambda: add_row("台北市", 0)).pack(side="right", padx=(0, 8))

        add_row("台南市", 100)
        card.after(50, redraw_bar)
        return rows

    def _apply_speed_preset(self, *_):
        """切換速度預設模式時自動填入建議值"""
        m = self.speed_preset.get()
        presets = {
            "快（偷懶模式）":   (0.15, 0.4,  0.5, 1.0, 0.8, 1.5, 20),
            "中（推薦）":         (0.3,  0.7,  0.8, 1.5, 1.5, 3.0, 30),
            "慢（高峰時段）":     (0.5,  1.2,  1.5, 2.5, 3.0, 5.0, 45),
            "極慢（嚴重塞車）":   (0.8,  2.0,  2.5, 4.0, 5.0, 8.0, 60),
        }
        if m in presets:
            a_lo, a_hi, p_lo, p_hi, b_lo, b_hi, to = presets[m]
            self.dly_act_lo.set(a_lo); self.dly_act_hi.set(a_hi)
            self.dly_page_lo.set(p_lo); self.dly_page_hi.set(p_hi)
            self.dly_btw_lo.set(b_lo);  self.dly_btw_hi.set(b_hi)
            self.dly_timeout.set(to)

    def _build_delay_config(self):
        return DelayConfig(
            action_lo=self.dly_act_lo.get(),  action_hi=self.dly_act_hi.get(),
            page_lo=self.dly_page_lo.get(),   page_hi=self.dly_page_hi.get(),
            between_lo=self.dly_btw_lo.get(), between_hi=self.dly_btw_hi.get(),
            wait_timeout=self.dly_timeout.get(),
        )

    def _update_counts(self, *_):
        try:
            tot = self.total_var.get()
        except Exception:
            tot = 0
        for block in (self.gender_pcts, self.nation_pcts, self.edu_pcts, self.orient_pcts, self.testing_pcts):
            for op, v, cnt in block:
                try: pv = v.get()
                except Exception: pv = 0
                cnt.set(f"{int(round(tot * pv / 100))} 筆")

    def log(self, msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        write_log_line(line)  # 同步寫入 number\log\log_*.txt
        try:
            tag = self._classify_log(msg)
            self.log_box.insert("end", line + "\n", tag)
            self.log_box.see("end")
        except Exception:
            tag = None
        # v1.0.53：嚴重紅色錯誤自動上報雲端（每 session 限一次，避免洗版）
        if tag == "error":
            try: self._maybe_report_error(msg)
            except Exception: pass

    def _maybe_report_error(self, trigger_msg):
        """v1.0.53/54：紅色 error log 觸發 → 把 log_box 內容打包 POST 到
        cloud_auth /report-error，由 Worker 存進 D1（並可選擇寄 email/Telegram）。
        v1.0.54 加機器識別資訊（hostname/win_user/mac/os_ver）。
        每 session 只發一次，避免錯誤連環觸發大量上報。
        靜默操作，不顯示任何 UI 彈窗。"""
        if getattr(self, "_error_reported", False):
            return
        self._error_reported = True
        try:
            log_text = self.log_box.get("1.0", "end")
        except Exception:
            log_text = trigger_msg or ""
        if len(log_text) > 60000:
            log_text = log_text[-60000:]

        # v1.0.54/56 機器識別資訊（共用 _machine_info）
        info = _machine_info()

        def _bg():
            import urllib.request, urllib.error
            try:
                body = json.dumps({
                    "version": VERSION,
                    "trigger": (trigger_msg or "")[:200],
                    "log_text": log_text,
                    **info,  # hostname/win_user/mac/os_ver
                }).encode("utf-8")
                req = urllib.request.Request(
                    CLOUD_AUTH_URL.rstrip("/") + "/report-error",
                    data=body,
                    headers={
                        "Content-Type": "application/json;charset=utf-8",
                        "User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)",
                    },
                    method="POST",
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    resp.read()
            except Exception:
                pass  # 靜默：上報失敗不影響主流程也不通知使用者
        threading.Thread(target=_bg, daemon=True).start()

    @staticmethod
    def _classify_log(msg):
        """依符號優先 → 關鍵詞次之 的策略分類 log 顏色。
           注意：詞彙比對需有「動作語氣」才染色，純名詞（如「失敗快照路徑」）不算。"""
        m = str(msg).strip()

        # ── 1. 強符號（決定性） ──
        # 紅
        for sym in ("✗", "✖", "❌"):
            if m.startswith(sym) or sym in m[:6]:
                return "error"
        # 綠
        for sym in ("✓", "✔"):
            if m.startswith(sym) or sym in m[:6]:
                return "success"
        # 黃
        for sym in ("⚠",):
            if sym in m:
                return "warn"

        # ── 2. 帶冒號的動作語氣 ──
        # 紅 — 「失敗：」「錯誤：」「異常：」等開頭明確說「結果是失敗」
        for kw in ("失敗：", "錯誤：", "異常：", "Exception", "Traceback", "ImportError"):
            if kw in m:
                return "error"
        # 綠
        for kw in ("成功：", "完成：", "已存", "已搬", "已開啟", "已推送", "已加入"):
            if kw in m:
                return "success"
        # 黃
        for kw in ("注意：", "警告：", "Warning", "繁忙", "等不到", "略過", "跳過", "請先"):
            if kw in m:
                return "warn"

        # ── 3. 強動作字（句中） ──
        if "💾" in m: return "success"
        if "▶ 開始" in m or "📦 已將" in m: return "success"
        if "■ 已停止" in m or "停止" == m or m.startswith("⏸"): return "warn"

        return "info"

    def status(self, txt):
        try: self.progress_var.set(txt)
        except Exception: pass

    def open_outdir(self):
        ensure_outdir()
        try:
            os.startfile(OUTPUT_DIR)
        except Exception as e:
            messagebox.showerror("開啟失敗", str(e))

    def open_logdir(self):
        ensure_outdir()
        p = os.path.join(OUTPUT_DIR, "log")
        try:
            os.startfile(p)
        except Exception as e:
            messagebox.showerror("開啟失敗", str(e))

    def open_debugdir(self):
        ensure_outdir()
        p = os.path.join(OUTPUT_DIR, "debug")
        try:
            os.startfile(p)
        except Exception as e:
            messagebox.showerror("開啟失敗", str(e))

    def clear_log(self):
        try:
            self.log_box.delete("1.0", "end")
        except Exception:
            pass

    def save_log(self):
        """v1.0.14：點下就直接存到 number/log/，不再問是否開啟"""
        try:
            content = self.log_box.get("1.0", "end-1c")
        except Exception:
            content = ""
        if not content.strip():
            self.log("⚠ 目前 log 是空的，沒有東西可儲存")
            return
        ensure_outdir()
        fp = os.path.join(OUTPUT_DIR, "log",
            f"log_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        try:
            with open(fp, "w", encoding="utf-8") as f:
                f.write(content)
            self.log(f"💾 已存 Log → {fp}")
        except Exception as e:
            self.log(f"⚠ Log 儲存失敗：{e}")

    def copy_log_clipboard(self):
        """v1.0.57：把當前 log box 全部內容複製到剪貼簿"""
        try:
            content = self.log_box.get("1.0", "end-1c")
        except Exception:
            content = ""
        if not content.strip():
            self.log("⚠ 目前 log 是空的，沒東西可複製")
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(content)
            self.root.update()  # 確保剪貼簿真的寫入
            self.log(f"📋 已複製 {len(content)} 字到剪貼簿")
        except Exception as e:
            self.log(f"⚠ 複製失敗：{e}")

    def show_log_full(self):
        """v1.0.57：開新 Toplevel 顯示完整 log（內含複製按鈕、可直接全選）"""
        try:
            content = self.log_box.get("1.0", "end-1c")
        except Exception:
            content = ""
        t = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
        win = tk.Toplevel(self.root)
        win.title(f"執行紀錄 — 完整 Log（{len(content)} 字）")
        win.geometry("1100x700")
        win.configure(bg=t["bg"])
        # top bar
        bar = tk.Frame(win, bg=t["bg"])
        bar.pack(fill="x", padx=12, pady=(10, 6))
        tk.Label(bar, text="🔍 完整執行紀錄", font=("微軟正黑體", 14, "bold"),
                 bg=t["bg"], fg=t["fg"]).pack(side="left")
        def _do_copy():
            try:
                win.clipboard_clear()
                win.clipboard_append(content)
                win.update()
                copy_btn.config(text="✓ 已複製")
                win.after(1500, lambda: copy_btn.config(text="📋 複製全部"))
            except Exception:
                pass
        copy_btn = tk.Button(bar, text="📋 複製全部", command=_do_copy,
                              font=("微軟正黑體", 10, "bold"),
                              bg=t["accent"], fg="#ffffff",
                              activebackground=t["accent_hover"], activeforeground="#ffffff",
                              relief="flat", bd=0, padx=14, pady=6, cursor="hand2")
        copy_btn.pack(side="right", padx=(8, 0))
        tk.Button(bar, text="關閉", command=win.destroy,
                   font=("微軟正黑體", 10),
                   bg=t["panel"], fg=t["fg"],
                   relief="flat", bd=0, padx=14, pady=6, cursor="hand2").pack(side="right")
        # log text area
        body_fr = tk.Frame(win, bg=t["bg"])
        body_fr.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        sb = tk.Scrollbar(body_fr, orient="vertical")
        sb.pack(side="right", fill="y")
        txt = tk.Text(body_fr, font=("Consolas", 11),
                       bg=t["log_bg"], fg=t["log_fg"],
                       insertbackground=t["log_fg"],
                       wrap="word", yscrollcommand=sb.set,
                       relief="flat", bd=0, padx=12, pady=10)
        txt.pack(side="left", fill="both", expand=True)
        sb.config(command=txt.yview)
        # 著色
        txt.tag_configure("success", foreground=t["log_success"])
        txt.tag_configure("warn",    foreground=t["log_warn"])
        txt.tag_configure("error",   foreground=t["log_error"], font=("Consolas", 11, "bold"))
        # 逐行 insert + 著色（沿用 _classify_log 邏輯）
        for line in content.split("\n"):
            tag = self._classify_log(line) if line else None
            if tag:
                txt.insert("end", line + "\n", tag)
            else:
                txt.insert("end", line + "\n")
        txt.see("end")
        # 預設選取全部方便 Ctrl+C 也能用
        txt.bind("<Control-a>", lambda e: (txt.tag_add("sel", "1.0", "end"), "break"))
        win.transient(self.root)
        win.lift()

    # ── 比例正規化驗證 ──
    def _validate_pcts(self, block, name):
        s = sum(v.get() for _, v, _ in block)
        if s != 100:
            messagebox.showerror("比例錯誤", f"{name} 合計 = {s}%，必須 = 100%")
            return False
        return True

    def _validate_city_rows(self, rows, name):
        if not rows:
            messagebox.showerror("比例錯誤", f"{name} 至少需 1 列")
            return False
        s = sum(p.get() for _, p, _ in rows)
        if s != 100:
            messagebox.showerror("比例錯誤", f"{name} 合計 = {s}%，必須 = 100%")
            return False
        return True

    def _build_profile_pool(self):
        """依比例展開成 N 筆 profile 列表（並隨機打亂）"""
        n = self.total_var.get()

        def expand(block):
            opts, ws = [], []
            for op, v, _ in block:
                opts.append(op); ws.append(v.get())
            return opts, ws

        def expand_city(rows):
            opts, ws = [], []
            for cv, pv, _ in rows:
                opts.append(cv.get()); ws.append(pv.get())
            return opts, ws

        g_opts, g_ws = expand(self.gender_pcts)
        n_opts, n_ws = expand(self.nation_pcts)
        e_opts, e_ws = expand(self.edu_pcts)
        o_opts, o_ws = expand(self.orient_pcts)
        t_opts, t_ws = expand(self.testing_pcts)
        r18_opts, r18_ws = expand_city(self.res18_rows)
        rC_opts, rC_ws  = expand_city(self.resCur_rows)

        ylo, yhi = sorted([self.year_lo.get(), self.year_hi.get()])

        pool = []
        for _ in range(n):
            pool.append({
                "gender": weighted_pick(g_opts, g_ws),
                "nation": weighted_pick(n_opts, n_ws),
                "year":   random.randint(ylo, yhi),
                "res18":  weighted_pick(r18_opts, r18_ws),
                "resCur": weighted_pick(rC_opts, rC_ws),
                "orient": weighted_pick(o_opts, o_ws),
                "edu":    weighted_pick(e_opts, e_ws),
                "testing": weighted_pick(t_opts, t_ws),
            })
        return pool

    # ── v1.0.29 增量儲存 Excel：完整 48 欄輸出 + 同日合併（依 header 對應） ──
    def _incremental_save_excel(self):
        if not self.results:
            return
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception as e:
            self.log(f"⚠ openpyxl 載入失敗，無法寫 Excel：{e}")
            return

        def _logical_blank(rec):
            """v1.0.34：parent=否 時清空子題（讓 xlsx 邏輯一致）"""
            r = dict(rec)
            if r.get("q6_std") == "否":
                for k in ("q6_hiv", "q6_hiv_reason", "q6_warts", "q6_syphilis",
                          "q6_gonorrhea", "q6_chlamydia", "q6_herpes",
                          "q6_hepA", "q6_hepC", "q6_other"):
                    r[k] = ""
            if r.get("q7_drug_use") == "否":
                for k in ("q7_amph", "q7_amph_method", "q7_ghb", "q7_mdma",
                          "q7_ketamine", "q7_rush", "q7_meph",
                          "q7_heroin", "q7_marijuana", "q7_other", "q7_status"):
                    r[k] = ""
            # Q6.1 HIV=否/空 → HIV 原因留空
            if r.get("q6_hiv") != "是":
                r["q6_hiv_reason"] = ""
            # 安非他命=否/空 → 使用方式留空
            if r.get("q7_amph") != "是":
                r["q7_amph_method"] = ""
            return r

        def _row_for_record(no, r):
            """組成一列：# / code / ts / 篩檢路徑 / 風險摘要 / 48 欄（先邏輯清空）"""
            r = _logical_blank(r)
            path, risk = _classify_record(r)
            row = [no, r.get("code", ""), r.get("ts", ""), path, risk]
            for k in COMPLETE_KEYS:
                row.append(r.get(k, ""))
            return row

        def _normalize_existing(old_rows, old_headers):
            """v1.0.35：以 header 名對應到新 schema（含摘要欄位置 +2 偏移）"""
            label_to_key = {v: k for k, v in OUTPUT_LABELS.items()}
            label_to_key.update({"性別": "gender", "國籍": "nation", "出生年": "year",
                                  "18歲前居住地": "res18", "現居住地": "resCur",
                                  "性傾向": "orient", "教育程度": "edu", "篩檢習慣": "testing_habit"})
            # 新 schema：[#, 代碼, 取得時間, 篩檢路徑, 風險摘要, ...COMPLETE_KEYS]
            col_to_output_pos = {}
            for i, h in enumerate(old_headers):
                if not h: continue
                hs = str(h).strip()
                if hs == "#":              col_to_output_pos[i] = 0
                elif hs == "諮詢代碼":     col_to_output_pos[i] = 1
                elif hs == "取得時間":     col_to_output_pos[i] = 2
                elif hs == "篩檢路徑":     col_to_output_pos[i] = 3
                elif hs == "風險摘要":     col_to_output_pos[i] = 4
                else:
                    key = label_to_key.get(hs)
                    if key and key in COMPLETE_KEYS:
                        col_to_output_pos[i] = 5 + COMPLETE_KEYS.index(key)
            normalized = []
            for old in old_rows:
                row = ["" for _ in range(len(OUTPUT_HEADERS))]
                for ci, val in enumerate(old):
                    pos = col_to_output_pos.get(ci)
                    if pos is not None:
                        row[pos] = val
                normalized.append(row)
            return normalized

        try:
            ensure_outdir()
            if not self._current_xlsx_path:
                # v1.0.53：固定檔名「諮詢代碼.xlsx」，多日多分頁；
                # 同日繼續取號 → 寫入今天日期的 sheet（合併既有列）
                # 跨日 → 新建今天日期的 sheet（其他日期 sheet 不動）
                # 沒檔案 → 全新檔
                xlsx_fname = "諮詢代碼.xlsx"
                xlsx_full = os.path.join(OUTPUT_DIR, xlsx_fname)
                today_sheet = datetime.datetime.now().strftime("%Y-%m-%d")
                self._existing_rows = []
                self._current_xlsx_path = xlsx_full
                if os.path.isfile(xlsx_full):
                    try:
                        wb_old = load_workbook(xlsx_full)
                        if today_sheet in wb_old.sheetnames:
                            ws_old = wb_old[today_sheet]
                            old_all = list(ws_old.iter_rows(values_only=True))
                            if old_all:
                                old_headers = list(old_all[0])
                                old_rows = [r for r in old_all[1:]
                                             if any(c not in (None, "") for c in r)]
                                self._existing_rows = _normalize_existing(old_rows, old_headers)
                            self.log(f"📁 諮詢代碼.xlsx 已有今天分頁（{len(self._existing_rows)} 筆），將合併寫入")
                        else:
                            self.log(f"📁 諮詢代碼.xlsx 已存在但今天 ({today_sheet}) 還沒分頁，將新增")
                    except Exception as e:
                        self.log(f"⚠ 讀取諮詢代碼.xlsx 失敗，改開新檔：{e}")
                        self._existing_rows = []
                else:
                    archive_old_outputs(self.log)
                    self.log("📁 諮詢代碼.xlsx 不存在，將建立新檔")
            # v1.0.53：固定檔多日多分頁；今天分頁永遠排在第一頁
            today_sheet = datetime.datetime.now().strftime("%Y-%m-%d")
            if os.path.isfile(self._current_xlsx_path):
                wb = load_workbook(self._current_xlsx_path)
                if today_sheet in wb.sheetnames:
                    del wb[today_sheet]
                ws = wb.create_sheet(today_sheet, 0)  # index 0 = 第一頁
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = today_sheet
            wb.active = 0
            ws.append(OUTPUT_HEADERS)
            # 收集每筆內容（既存 + 新）並依路徑分流
            all_rows = []           # [(content, path)]
            row_idx = 0
            # 既存合併：判路徑
            for old in (self._existing_rows or []):
                row_idx += 1
                content = list(old)
                while len(content) < len(OUTPUT_HEADERS):
                    content.append("")
                content[0] = row_idx
                content = content[:len(OUTPUT_HEADERS)]
                # 從 row 內容反推 record dict 來算 path（取需要的 key）
                # 簡化：直接看 content[3] 若有值用之，否則重算
                rec = {COMPLETE_KEYS[i]: content[5 + i] for i in range(len(COMPLETE_KEYS))}
                if not content[3]:
                    p, ri = _classify_record(rec)
                    content[3] = p; content[4] = ri
                path = content[3] or "簡單"
                all_rows.append((content, path))
            # 新筆
            for r in self.results:
                row_idx += 1
                content = _row_for_record(row_idx, r)
                path = content[3] or "簡單"
                all_rows.append((content, path))
            # 寫到主分頁 + 套色
            FILL_BY_PATH = {
                "簡單":   None,
                "性行為": PatternFill("solid", fgColor="FFE0B2"),  # 淡橘
                "感染史": PatternFill("solid", fgColor="FFCDD2"),  # 淡紅
                "藥物史": PatternFill("solid", fgColor="FFCDD2"),  # 淡紅
            }
            for content, path in all_rows:
                ws.append(content)
                row_n = ws.max_row
                fill = FILL_BY_PATH.get(path)
                if fill:
                    for col in range(1, len(OUTPUT_HEADERS) + 1):
                        ws.cell(row=row_n, column=col).fill = fill
            # 標題列樣式
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF", size=10)
                c.fill = PatternFill("solid", fgColor="1565C0")
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 8
            for col_idx in range(6, len(OUTPUT_HEADERS) + 1):
                col_letter = ws.cell(1, col_idx).column_letter
                ws.column_dimensions[col_letter].width = 16
            ws.freeze_panes = "F2"  # 凍結 # / 代碼 / 時間 / 路徑 / 摘要 + 標題列
            # v1.0.52：移除原本依風險路徑拆 sheet 的邏輯（簡單/性行為/感染史/藥物史）
            #         只留單一日期分頁，整體更乾淨；列上仍會依路徑套淡色背景作區分
            self._save_xlsx_atomic(wb, self._current_xlsx_path)
        except Exception as e:
            self.log(f"⚠ 寫 Excel 失敗：{e}")

    # ── v1.0.38 原子寫入 + 鎖檔影子檔 ────────────────────
    def _save_xlsx_atomic(self, wb, target_path):
        """先寫 .tmp 再 os.replace 到 target；若 target 被開啟（PermissionError）
        則改寫 _LIVE.xlsx 影子檔（同樣有完整資料）。下次寫主檔成功後
        會自動清掉影子檔。內容永遠是當下累積結果，重試不會遺失。"""
        tmp_path = target_path + ".writing.tmp"
        try:
            wb.save(tmp_path)
        except Exception as e:
            self.log(f"⚠ 寫暫存檔失敗：{e}")
            return False

        # 嘗試原子換掉主檔
        try:
            os.replace(tmp_path, target_path)
        except PermissionError:
            # 主檔被開啟，改寫 _LIVE 影子檔
            stem, ext = os.path.splitext(target_path)
            live_path = f"{stem}_LIVE{ext}"
            try:
                os.replace(tmp_path, live_path)
                if not getattr(self, "_warned_locked", False):
                    self.log(
                        f"⚠ 主檔被開啟中，已寫入影子檔 "
                        f"{os.path.basename(live_path)}（資料完整、不會遺失）"
                    )
                    self.log("   請關閉主檔後，下一筆會自動回寫主檔。")
                    self._warned_locked = True
            except Exception as e:
                self.log(f"⚠ 連影子檔都寫不進去：{e}")
                # 最後保險：留住 .tmp 不刪
                return False
            return False
        except Exception as e:
            self.log(f"⚠ 換檔失敗（將下次重試）：{e}")
            return False

        # 主檔成功 → 清掉之前的影子檔（若存在）
        stem, ext = os.path.splitext(target_path)
        live_path = f"{stem}_LIVE{ext}"
        if os.path.isfile(live_path):
            try:
                os.remove(live_path)
                self.log("✓ 主檔已可寫入，影子檔自動移除")
            except Exception:
                pass
        if getattr(self, "_warned_locked", False):
            self._warned_locked = False
        return True

    # ── 啟動 ──
    # ── v1.0.21 模式切換 + xlsx 匯入匯出 ──
    def _on_mode_change(self):
        is_complete = (self.mode_var.get() == "完整（匯入xlsx）")
        try:
            if is_complete:
                self._batch_xlsx_tools.pack(side="left", padx=4)  # v1.0.44
                self._batch_topbar.pack_forget()
                self._batch_midbar.pack_forget()
                self._batch_preview_fr.pack(fill="x", padx=6, pady=4)
                self.log("📋 切換到完整模式（請匯入 xlsx）")
            else:
                self._batch_xlsx_tools.pack_forget()  # v1.0.44
                self._batch_preview_fr.pack_forget()
                self._batch_topbar.pack(fill="x", padx=6, pady=4)
                self._batch_midbar.pack(fill="x", padx=6, pady=4)
                self.log("📊 切換到簡易模式")
        except Exception as e:
            self.log(f"⚠ 模式切換異常：{e}")

    def _export_sample_xlsx(self):
        ensure_outdir()
        fp = os.path.join(OUTPUT_DIR, "sample_complete_mode.xlsx")
        try:
            make_sample_xlsx(fp)
            self.log(f"💾 範例 xlsx 已產生：{fp}")
            try: os.startfile(fp)
            except Exception: pass
        except Exception as e:
            messagebox.showerror("匯出失敗", str(e))

    def _import_xlsx(self):
        from tkinter import filedialog
        fp = filedialog.askopenfilename(
            title="選擇要匯入的 xlsx",
            filetypes=[("Excel 檔", "*.xlsx"), ("所有檔案", "*.*")],
            initialdir=OUTPUT_DIR,
        )
        if not fp: return
        try:
            profiles, warn = import_xlsx_profiles(fp)
            if isinstance(warn, dict) and warn.get("fatal"):
                messagebox.showerror("匯入錯誤", warn["fatal"]); return
            if not profiles:
                messagebox.showwarning("匯入警告", "沒有讀到任何資料列")
                return
            # v1.0.33：發現 blank 或 invalid 時，跳出驗證 modal 讓使用者決定
            blanks = warn.get("blank", []) if isinstance(warn, dict) else []
            invalids = warn.get("invalid", []) if isinstance(warn, dict) else []
            if blanks or invalids:
                go = self._show_validation_modal(fp, profiles, blanks, invalids)
                if not go:
                    self.log(f"📋 已取消匯入（使用者選擇回去修 {os.path.basename(fp)}）")
                    return
            # 通過驗證或使用者選擇繼續
            self._imported_profiles = profiles
            self.import_status_var.set(f"✓ 已匯入 {len(profiles)} 筆")
            self.log(f"📥 已匯入 {len(profiles)} 筆 → {os.path.basename(fp)}")
            for ridx, fields in blanks[:5]:
                self.log(f"  ⚠ 列 {ridx} 未填：{', '.join(fields[:6])}{'...' if len(fields)>6 else ''}（已用預設值）")
            if len(blanks) > 5:
                self.log(f"  ⚠ 另有 {len(blanks)-5} 列有未填欄位")
            for ridx, label, bad, dv in invalids[:5]:
                self.log(f"  ⚠ 列 {ridx} {label}=「{bad}」不合法，已改為「{dv}」")
            if self.mode_var.get() != "完整（匯入xlsx）":
                self.mode_var.set("完整（匯入xlsx）")
                self._on_mode_change()
            self._refresh_preview_tree()
        except Exception as e:
            messagebox.showerror("匯入失敗", str(e))

    def _show_validation_modal(self, file_path, profiles, blanks, invalids):
        """v1.0.33：xlsx 匯入驗證警示 modal，回 True=繼續匯入，False=取消"""
        t = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
        modal = tk.Toplevel(self.root)
        modal.title("匯入資料驗證")
        modal.configure(bg=t["bg"])
        modal.geometry("760x560")
        modal.transient(self.root)
        modal.grab_set()
        modal.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 760) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 560) // 2
        modal.geometry(f"+{x}+{y}")
        # 標題
        tk.Label(modal, text="⚠ xlsx 匯入發現問題", bg=t["bg"], fg=t["warn"],
                  font=("微軟正黑體", 16, "bold")).pack(pady=(20, 6))
        tk.Label(modal,
                 text=f"檔案：{os.path.basename(file_path)}　／　共 {len(profiles)} 筆資料\n"
                      f"未填欄位：{len(blanks)} 列　／　無效值：{len(invalids)} 處",
                 bg=t["bg"], fg=t["fg"], font=("微軟正黑體", 11)).pack(pady=4)
        # 詳細列表
        list_fr = ttk.Frame(modal); list_fr.pack(fill="both", expand=True, padx=20, pady=8)
        text = tk.Text(list_fr, bg=t["log_bg"], fg=t["log_fg"], font=("Consolas", 10),
                       wrap="word", height=20)
        text.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(list_fr, orient="vertical", command=text.yview)
        sb.pack(side="right", fill="y")
        text.configure(yscrollcommand=sb.set)
        text.tag_configure("blank", foreground="#ffb74d")
        text.tag_configure("invalid", foreground="#ef5350")
        text.tag_configure("hint", foreground="#90a4ae")
        if blanks:
            text.insert("end", "▼ 未填欄位（將套用預設值）：\n", "blank")
            for ridx, fields in blanks[:80]:
                text.insert("end", f"  列 {ridx}：{', '.join(fields)}\n")
            if len(blanks) > 80:
                text.insert("end", f"  ... 另有 {len(blanks)-80} 列\n", "hint")
            text.insert("end", "\n")
        if invalids:
            text.insert("end", "▼ 無效值（將改為預設）：\n", "invalid")
            for ridx, label, bad, dv in invalids[:80]:
                text.insert("end", f"  列 {ridx} {label}=「{bad}」 → 改為「{dv}」\n")
            if len(invalids) > 80:
                text.insert("end", f"  ... 另有 {len(invalids)-80} 處\n", "hint")
            text.insert("end", "\n")
        text.insert("end",
            "選擇：\n"
            "  ✅ 用預設值繼續：照上面預設值套用後直接執行取號（缺漏處用預設）\n"
            "  📝 取消，去修 xlsx：先回去把空格補齊，再重新匯入\n",
            "hint")
        text.configure(state="disabled")
        # 按鈕
        result = {"go": False}
        btn_fr = tk.Frame(modal, bg=t["bg"])
        btn_fr.pack(pady=14)
        def do_continue():
            result["go"] = True
            modal.destroy()
        def do_cancel():
            result["go"] = False
            modal.destroy()
        tk.Button(btn_fr, text="✅ 用預設值繼續", bg=t["accent"], fg="white",
                  font=("微軟正黑體", 12, "bold"), padx=20, pady=10, bd=0, cursor="hand2",
                  command=do_continue).pack(side="left", padx=10)
        tk.Button(btn_fr, text="📝 取消，去修 xlsx", bg="#90a4ae", fg="white",
                  font=("微軟正黑體", 12, "bold"), padx=20, pady=10, bd=0, cursor="hand2",
                  command=do_cancel).pack(side="left", padx=10)
        modal.wait_window()
        return result["go"]

    def _refresh_preview_tree(self):
        """v1.0.23：把匯入的 profiles 顯示在預覽 Treeview"""
        try:
            tree = self.preview_tree
        except AttributeError:
            return
        for item in tree.get_children():
            tree.delete(item)
        for i, p in enumerate(self._imported_profiles, 1):
            tree.insert("", "end", values=(
                i, p.get("gender", ""), p.get("year", ""),
                p.get("res18", ""), p.get("resCur", ""),
                p.get("orient", ""), p.get("edu", ""),
                p.get("q1_sex", ""), p.get("q6_std", ""),
                p.get("q7_drug_use", ""), p.get("testing_habit", ""),
            ))

    # ── v1.0.21 取號中設定鎖定 ──
    def _set_settings_locked(self, locked):
        """取號開始 → True（全部 disable）；停止後 → False"""
        # 簡單做法：遞迴掃過 root 的所有 ttk.Entry / Spinbox / Combobox / Radiobutton / Checkbutton
        def walk(w):
            try:
                cls = w.winfo_class()
                if cls in ("TEntry", "Entry", "TSpinbox", "Spinbox",
                           "TCombobox", "TRadiobutton", "Radiobutton",
                           "TCheckbutton", "Checkbutton"):
                    # 不鎖控制按鈕（start/stop/pause）— 它們是 TButton
                    state = "disabled" if locked else "!disabled"
                    try: w.state([state])
                    except Exception:
                        try: w.configure(state="disabled" if locked else "normal")
                        except Exception: pass
            except Exception: pass
            for c in w.winfo_children():
                walk(c)
        walk(self.root)

    # ── v1.0.21 續傳機制 ──
    def _save_resume_state(self):
        """停止關閉前保存：未完成的 pool + 已完成 index"""
        try:
            if not self.thread or not self.thread.is_alive():
                return
            state = {
                "version": VERSION,
                "saved_at": datetime.datetime.now().isoformat(),
                "results_count": len(self.results),
                "pool": getattr(self, "_current_pool", []),
                "completed_index": getattr(self, "_completed_index", 0),
                "xlsx_path": self._current_xlsx_path,
                "mode": self.mode_var.get(),
            }
            ensure_outdir()
            rs_path = os.path.join(OUTPUT_DIR, "resume_state.json")
            with open(rs_path, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False, indent=2, default=str)
            _hide_path(rs_path)  # v1.0.43 L4：與 settings.json 一致
        except Exception:
            pass

    def _check_resume(self):
        """啟動時檢查是否有未完成的進度"""
        rs_path = os.path.join(OUTPUT_DIR, "resume_state.json")
        if not os.path.exists(rs_path):
            return
        try:
            with open(rs_path, "r", encoding="utf-8") as f:
                state = json.load(f)
        except Exception:
            return
        pool = state.get("pool", [])
        done = int(state.get("completed_index", 0))
        if not pool or done >= len(pool):
            try: os.remove(rs_path)
            except Exception: pass
            return
        remaining = len(pool) - done
        # 跳出 modal
        self.root.after(300, lambda: self._show_resume_modal(state, pool, done, remaining, rs_path))

    def _show_resume_modal(self, state, pool, done, remaining, rs_path):
        t = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
        modal = tk.Toplevel(self.root)
        modal.title("續傳上次進度？")
        modal.configure(bg=t["bg"])
        modal.geometry("520x300")
        modal.transient(self.root)
        modal.grab_set()
        # 居中
        modal.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 520) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 300) // 2
        modal.geometry(f"+{x}+{y}")
        # 標題
        tk.Label(modal, text="🔄 偵測到上次未完成的取號", bg=t["bg"], fg=t["accent"],
                  font=("微軟正黑體", 16, "bold")).pack(pady=(20, 10))
        info = (f"上次跑到第 {done}/{len(pool)} 筆\n"
                f"剩餘 {remaining} 筆\n"
                f"上次儲存時間：{state.get('saved_at', '?')[:19]}\n"
                f"上次 Excel：{os.path.basename(state.get('xlsx_path') or '—')}")
        tk.Label(modal, text=info, bg=t["bg"], fg=t["fg"],
                  font=("微軟正黑體", 11), justify="center").pack(pady=10)
        btn_fr = tk.Frame(modal, bg=t["bg"])
        btn_fr.pack(pady=20)
        def do_resume():
            modal.destroy()
            self._imported_profiles = pool[done:]  # 剩餘
            self.mode_var.set("完整（匯入xlsx）")
            self._on_mode_change()
            self.import_status_var.set(f"✓ 續傳剩 {remaining} 筆")
            self.log(f"🔄 續傳：將從第 {done+1} 筆開始（剩 {remaining} 筆）")
            try:
                # 接續寫入舊 Excel
                self._current_xlsx_path = state.get("xlsx_path")
            except Exception: pass
            try: os.remove(rs_path)
            except Exception: pass
        def do_fresh():
            modal.destroy()
            try: os.remove(rs_path)
            except Exception: pass
            self.log("🆕 已選擇重新開始（清除上次進度）")
        tk.Button(btn_fr, text=f"📂 續傳剩 {remaining} 筆", bg=t["accent"], fg="white",
                  font=("微軟正黑體", 12, "bold"), padx=20, pady=10,
                  command=do_resume, bd=0, cursor="hand2").pack(side="left", padx=10)
        tk.Button(btn_fr, text="🆕 重新開始", bg="#90a4ae", fg="white",
                  font=("微軟正黑體", 12, "bold"), padx=20, pady=10,
                  command=do_fresh, bd=0, cursor="hand2").pack(side="left", padx=10)

    # ── v1.0.20 數字輸入框（取代 Spinbox 上下箭頭） ──
    def _mk_int_entry(self, parent, var, width=6, justify="center"):
        """整數輸入框：只接受 0-9"""
        def vc(P):
            return P == "" or P.isdigit() or (P[0] == "-" and P[1:].isdigit())
        e = ttk.Entry(parent, textvariable=var, width=width, justify=justify,
                      validate="key", validatecommand=(parent.register(vc), "%P"))
        return e

    def _mk_float_entry(self, parent, var, width=6, justify="center"):
        """小數輸入框：接受 0-9 與 . """
        def vc(P):
            if P == "": return True
            try: float(P); return True
            except Exception: return False
        e = ttk.Entry(parent, textvariable=var, width=width, justify=justify,
                      validate="key", validatecommand=(parent.register(vc), "%P"))
        return e

    # ── v1.0.18 主題切換 + 動畫 ──
    def _apply_theme(self, name):
        t = THEMES.get(name, THEMES[DEFAULT_THEME])
        try:
            self.root.configure(bg=t["bg"])
            style = ttk.Style()
            try: style.theme_use("clam")
            except Exception: pass
            # 通用元件
            style.configure(".", background=t["bg"], foreground=t["fg"])
            style.configure("TFrame", background=t["bg"])
            style.configure("TLabel", background=t["bg"], foreground=t["fg"])
            style.configure("TLabelframe", background=t["bg"], foreground=t["accent"], borderwidth=1)
            style.configure("TLabelframe.Label", background=t["bg"], foreground=t["accent"],
                           font=("微軟正黑體", 10, "bold"))
            style.configure("TButton", background=t["panel"], foreground=t["fg"],
                           padding=(8, 4), borderwidth=1)
            style.map("TButton",
                background=[("active", t["accent"]), ("pressed", t["accent_hover"])],
                foreground=[("active", "white"), ("pressed", "white")])
            # 大按鈕（開始 / 暫停 / 停止）
            style.configure("Big.TButton", background=t["accent"], foreground="white",
                           font=("微軟正黑體", 14, "bold"), padding=(20, 10))
            style.map("Big.TButton",
                background=[("active", t["accent_hover"]), ("disabled", "#90a4ae")])
            style.configure("BigPause.TButton", background=t["warn"], foreground="white",
                           font=("微軟正黑體", 14, "bold"), padding=(16, 10))
            style.map("BigPause.TButton",
                background=[("active", "#bf360c"), ("disabled", "#bdbdbd")])
            style.configure("BigStop.TButton", background=t["error"], foreground="white",
                           font=("微軟正黑體", 14, "bold"), padding=(16, 10))
            style.map("BigStop.TButton",
                background=[("active", "#7f0000"), ("disabled", "#bdbdbd")])
            # Entry / Spinbox / Combobox
            style.configure("TEntry", fieldbackground=t["panel"], foreground=t["fg"], borderwidth=1)
            style.configure("TSpinbox", fieldbackground=t["panel"], foreground=t["fg"])
            style.configure("TCombobox", fieldbackground=t["panel"], foreground=t["fg"])
            # Progressbar (legacy ttk - 雖然主進度條改 Canvas，這裡保留設定以防其他地方用)
            style.configure("TProgressbar", troughcolor=t["panel"], background=t["accent"],
                           borderwidth=0, thickness=18)
            # 自訂 Canvas 進度條（v1.0.48 跑者衣服色也跟主題走）
            try: self.pb.configure_theme(t["panel"], t["accent"], t["accent_hover"])
            except Exception: pass
            # Checkbutton
            style.configure("TCheckbutton", background=t["bg"], foreground=t["fg"])
            # log box
            try:
                self.log_box.configure(bg=t["log_bg"], fg=t["log_fg"], insertbackground=t["log_fg"])
                self.log_box.tag_configure("success", foreground=t["log_success"])
                self.log_box.tag_configure("warn",    foreground=t["log_warn"])
                self.log_box.tag_configure("error",   foreground=t["log_error"], font=("Consolas", 10, "bold"))
                self.log_box.tag_configure("info",    foreground=t["log_fg"])
            except Exception:
                pass
            self.log(f"🎨 已套用主題：{name}")
        except Exception as e:
            self.log(f"⚠ 主題套用失敗：{e}")

    def _start_spinner(self):
        self._spinner_running = True
        self._tick_spinner()

    def _stop_spinner(self):
        self._spinner_running = False
        try: self.spinner_label.configure(text="●", foreground="#90a4ae")
        except Exception: pass

    def _tick_spinner(self):
        if not self._spinner_running:
            return
        frames = ["⣷", "⣯", "⣟", "⡿", "⢿", "⣻", "⣽", "⣾"]
        try:
            t = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
            self.spinner_label.configure(text=frames[self._spinner_idx % len(frames)],
                                          foreground=t["accent"])
        except Exception:
            pass
        self._spinner_idx += 1
        self.root.after(120, self._tick_spinner)

    def _flash_stats(self, color="#4caf50"):
        """成功取得一筆 → 統計列短暫閃爍綠色背景"""
        try:
            orig_bg = self.stats_var
            # 透過 Label 配置直接改 fg 模擬閃爍
            label = None
            for child in self.root.winfo_children():
                # 找 stats label，太繁瑣 — 改用更簡單方法：用 progress label 文字色閃爍
                pass
            # 簡化版：用 spinner_label 暫時亮綠
            t = THEMES.get(self.theme_var.get(), THEMES[DEFAULT_THEME])
            self.spinner_label.configure(foreground=t["success"], text="✓")
            self.root.after(200, lambda: self.spinner_label.configure(text="●") if self._spinner_running else None)
            self.root.after(220, lambda: self._tick_spinner() if self._spinner_running else None)
        except Exception:
            pass

    # ── v1.0.15 輸出路徑控制 ──
    def _pick_output_dir(self):
        from tkinter import filedialog
        cur = self.output_dir_var.get() or DEFAULT_OUTPUT_DIR
        if not os.path.isdir(cur):
            cur = os.path.dirname(cur) if os.path.dirname(cur) else DEFAULT_OUTPUT_DIR
        new = filedialog.askdirectory(title="選擇輸出資料夾", initialdir=cur)
        if new:
            self.output_dir_var.set(os.path.normpath(new))
            self._apply_output_dir()

    def _apply_output_dir(self):
        new_path = (self.output_dir_var.get() or "").strip()
        if not new_path:
            messagebox.showwarning("路徑錯誤", "輸出路徑不能空白"); return
        try:
            set_output_dir(new_path)  # 自動建資料夾 + 子目錄
            self.log(f"📁 輸出資料夾已切換：{new_path}")
            self._save_settings()
        except Exception as e:
            messagebox.showerror("套用失敗", f"無法建立或寫入：\n{new_path}\n\n{e}")

    def _reset_output_dir(self):
        self.output_dir_var.set(DEFAULT_OUTPUT_DIR)
        self._apply_output_dir()

    def _reset_stats(self):
        try:
            self.stats_var.set("平均 — 秒/筆   |   已用 00m00s   |   剩餘 —   |   預計完成 —")
            self.pb.configure_value(0, 100)
        except Exception: pass

    def _build_pool_for_mode(self):
        """v1.0.21：依當前模式產 pool"""
        if self.mode_var.get() == "完整（匯入xlsx）":
            if not self._imported_profiles:
                messagebox.showerror("無資料", "完整模式請先匯入 xlsx")
                return None
            return list(self._imported_profiles)
        else:
            return self._build_profile_pool()

    def _smart_start(self):
        """v1.0.26：依當前分頁智慧分派 — 批次→start()；單筆→start_single()"""
        try:
            cur = self.notebook.index(self.notebook.select())
        except Exception:
            cur = 0
        if cur == 1:
            self.start_single()
        else:
            self.start()
        # 依分頁更新按鈕文字
        self._update_start_btn_label()

    def _update_start_btn_label(self):
        """依當前分頁更新「▶ 開始」按鈕文字"""
        try:
            cur = self.notebook.index(self.notebook.select())
            if cur == 1:
                self.start_btn.configure(text="▶ 開始單筆取號")
            else:
                self.start_btn.configure(text="▶ 開始批次取號")
        except Exception:
            pass

    def start(self):
        # 驗證
        for blk, name in [
            (self.gender_pcts, "性別"),
            (self.nation_pcts, "國籍"),
            (self.edu_pcts, "教育程度"),
            (self.orient_pcts, "性傾向"),
            (self.testing_pcts, "篩檢習慣"),
        ]:
            if not self._validate_pcts(blk, name):
                return
        if not self._validate_city_rows(self.res18_rows, "18歲前居住地"): return
        if not self._validate_city_rows(self.resCur_rows, "現居住地"): return
        if self.total_var.get() < 1:
            messagebox.showerror("錯誤", "總筆數至少 1")
            return

        ensure_outdir()
        self._save_settings()
        pool = self._build_pool_for_mode()
        if pool is None:
            return
        self.results = []
        # v1.0.21 續傳時保留 xlsx 路徑
        if not self._current_xlsx_path:
            self._current_xlsx_path = None
        self._current_pool = pool
        self._completed_index = 0
        self.stop_evt.clear()
        self.pause_evt.clear()
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.pause_btn.config(state="normal", text="⏸ 暫停")
        self.pb.configure_value(0, len(pool))
        self._reset_stats()
        self._start_spinner()
        self._set_settings_locked(True)  # v1.0.21 反白
        self.thread = threading.Thread(target=self._run, args=(pool,), daemon=True)
        self.thread.start()

    # ── v1.0.19/25 單筆取號 ──
    def start_single(self):
        if self.thread and self.thread.is_alive():
            messagebox.showwarning("執行中", "目前批次取號還在跑，請先停止再用單筆模式")
            return
        ensure_outdir()
        self._save_settings()
        sv = self.single_vars
        is_complete_single = (self.single_mode_var.get() == "完整")
        if is_complete_single:
            # v1.0.25 完整模式：收齊所有 48 欄
            prof = {}
            for k, _, dv, _ in COMPLETE_FIELDS:
                if k in sv:
                    val = sv[k].get()
                    prof[k] = val if val != "" else dv
                else:
                    prof[k] = dv
            try: prof["year"] = int(prof.get("year") or 1990)
            except Exception: prof["year"] = 1990
            # 兼容舊 worker 路徑（is_complete 偵測用 q1_sex 存在）
        else:
            prof = {
                "gender":  sv["gender"].get(),
                "nation":  sv["nation"].get(),
                "year":    int(sv["year"].get()),
                "res18":   sv["res18"].get(),
                "resCur":  sv["resCur"].get(),
                "orient":  sv["orient"].get(),
                "edu":     sv["edu"].get(),
                "testing": sv["testing"].get(),
            }
        self.results = []
        self._current_xlsx_path = None
        self.stop_evt.clear()
        self.pause_evt.clear()
        self.start_btn.config(state="disabled")
        self.pause_btn.config(state="disabled")
        self.stop_btn.config(state="disabled")
        self.pb.configure_value(0, 1)
        self._reset_stats()
        self._start_spinner()
        self.single_result_var.set("執行中…")
        self.thread = threading.Thread(target=self._run, args=([prof],), daemon=True)
        self.thread.start()

    def stop(self):
        self.stop_evt.set()
        # 萬一還在暫停狀態，先解除暫停讓主迴圈跑完當前 → 進結束流程
        if self.pause_evt.is_set():
            self.pause_evt.clear()
        self.log("■ 收到停止訊號，等本筆結束 → 存檔 → 自動關閉…")
        try:
            self.stop_btn.config(state="disabled")
            self.pause_btn.config(state="disabled")
        except Exception:
            pass
        # v1.0.51 跑者切「喘氣」狀態
        try: self.pb.set_state('panting')
        except Exception: pass

    def toggle_pause(self):
        """暫停/繼續：每筆任務開頭/中段都會檢查 pause_evt，被 set 即進入忙等迴圈"""
        if self.pause_evt.is_set():
            self.pause_evt.clear()
            self.pause_btn.config(text="⏸ 暫停")
            self.log("▶ 已繼續")
            # v1.0.51 清掉跑者喘氣 overlay，讓奔跑動畫接手
            try: self.pb.set_state('idle')
            except Exception: pass
        else:
            self.pause_evt.set()
            self.pause_btn.config(text="▶ 繼續")
            self.log("⏸ 已暫停（按「繼續」恢復）")
            try: self.pb.set_state('panting')
            except Exception: pass

    # ── v1.0.12 設定持久化 ──
    def _collect_settings(self):
        d = {
            "output_dir": self.output_dir_var.get(),  # v1.0.15
            "theme": self.theme_var.get(),  # v1.0.18
            "total": self.total_var.get(),
            "year_lo": self.year_lo.get(),
            "year_hi": self.year_hi.get(),
            "speed_preset": self.speed_preset.get(),
            "dly_act_lo":  self.dly_act_lo.get(),  "dly_act_hi":  self.dly_act_hi.get(),
            "dly_page_lo": self.dly_page_lo.get(), "dly_page_hi": self.dly_page_hi.get(),
            "dly_btw_lo":  self.dly_btw_lo.get(),  "dly_btw_hi":  self.dly_btw_hi.get(),
            "dly_timeout": self.dly_timeout.get(),
            "gender":  [(op, v.get()) for op, v, _ in self.gender_pcts],
            "nation":  [(op, v.get()) for op, v, _ in self.nation_pcts],
            "edu":     [(op, v.get()) for op, v, _ in self.edu_pcts],
            "orient":  [(op, v.get()) for op, v, _ in self.orient_pcts],
            "testing": [(op, v.get()) for op, v, _ in self.testing_pcts],
            "res18":   [(cv.get(), pv.get()) for cv, pv, _ in self.res18_rows],
            "resCur":  [(cv.get(), pv.get()) for cv, pv, _ in self.resCur_rows],
        }
        return d

    def _save_settings(self):
        try:
            ensure_outdir()
            existing = {}
            if os.path.exists(SETTINGS_FILE):
                try:
                    with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                        existing = json.load(f)
                except Exception:
                    existing = {}
            merged = {**existing, **self._collect_settings()}
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(merged, f, ensure_ascii=False, indent=2)
            _hide_path(SETTINGS_FILE)  # v1.0.42 隱藏 settings.json
        except Exception as e:
            self.log(f"⚠ 設定存檔失敗：{e}")

    def _load_settings(self):
        if not os.path.exists(SETTINGS_FILE):
            return
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
        except Exception as e:
            self.log(f"⚠ 設定讀取失敗：{e}")
            return
        # 暫停 trace 避免 set 過程被 _balance 干擾
        self._balancing = True
        try:
            # v1.0.15 輸出路徑（先還原才能讓後續存檔走正確位置）
            if "output_dir" in d and d["output_dir"]:
                try:
                    self.output_dir_var.set(d["output_dir"])
                    set_output_dir(d["output_dir"])
                except Exception as e:
                    self.log(f"⚠ 還原輸出路徑失敗：{e}")
            # v1.0.18 主題
            if "theme" in d and d["theme"] in THEMES:
                self.theme_var.set(d["theme"])
                self._apply_theme(d["theme"])
            if "total" in d:    self.total_var.set(d["total"])
            if "year_lo" in d:  self.year_lo.set(d["year_lo"])
            if "year_hi" in d:  self.year_hi.set(d["year_hi"])
            if "speed_preset" in d: self.speed_preset.set(d["speed_preset"])
            for k_ui, k_dict in [
                ("dly_act_lo","dly_act_lo"), ("dly_act_hi","dly_act_hi"),
                ("dly_page_lo","dly_page_lo"), ("dly_page_hi","dly_page_hi"),
                ("dly_btw_lo","dly_btw_lo"), ("dly_btw_hi","dly_btw_hi"),
                ("dly_timeout","dly_timeout"),
            ]:
                if k_dict in d: getattr(self, k_ui).set(d[k_dict])
            def _restore_pct(saved, block):
                if not saved: return
                m = dict(saved)
                for op, v, _ in block:
                    if op in m: v.set(m[op])
            _restore_pct(d.get("gender"),  self.gender_pcts)
            _restore_pct(d.get("nation"),  self.nation_pcts)
            _restore_pct(d.get("edu"),     self.edu_pcts)
            _restore_pct(d.get("orient"),  self.orient_pcts)
            _restore_pct(d.get("testing"), self.testing_pcts)
            # 居住地：移除預設那行 + 從 settings 重建
            def _restore_city(saved, rows, add_row_cb):
                if not saved: return
                # 清空現有列
                for cv, pv, fr in list(rows):
                    fr.destroy()
                rows.clear()
                for city, pct in saved:
                    add_row_cb(city, int(pct))
            # 取出對應的 add_row 函式（reflect from frame）— 我們需先儲存 add_row callback
            # 用替代法：直接設定第一列、加入剩餘列
            if d.get("res18"):
                # 簡化：直接把 row[0] 改值，其餘呼叫該 block 既有的 +添加
                pass  # 預設行為夠用（首次有預設 1 列台南市 100%）
        finally:
            self._balancing = False
        self._update_counts()
        self.log("📁 已載入上次設定")

    def _on_close(self):
        try:
            self._save_settings()
        except Exception:
            pass
        self.root.destroy()

    def _run(self, pool):
        self.log(f"▶ 開始批次取號，共 {len(pool)} 筆")
        dly = self._build_delay_config()
        self.log(f"   延遲設定：動作 {dly.action_lo}~{dly.action_hi}s / 頁面 {dly.page_lo}~{dly.page_hi}s "
                 f"/ 每筆間隔 {dly.between_lo}~{dly.between_hi}s / 逾時 {dly.wait_timeout}s")
        # v1.0.14 統計起始時間
        run_start_ts = time.time()
        self._batch_start_ts = run_start_ts  # v1.0.59 給 _finish 算 duration 用
        self._batch_target = len(pool)        # v1.0.59 給 status 判斷用
        completed = 0
        worker = HivaWorker(self.log, self.status, lambda code: None, self.stop_evt, delay_cfg=dly)
        if not worker.start_browser():
            self._finish()
            return
        try:
            for i, prof in enumerate(pool, 1):
                # v1.0.12 暫停檢查（在每筆開頭）
                while self.pause_evt.is_set() and not self.stop_evt.is_set():
                    time.sleep(0.5)
                if self.stop_evt.is_set():
                    self.log("■ 已停止")
                    break
                self.status(f"取第 {i}/{len(pool)} 筆 — {prof['gender']}/{prof['year']}/{prof['edu']}")
                code = worker.fetch_one(prof)
                ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if code:
                    self.log(f"✓ 第 {i} 筆 → {code}  ({prof['gender']} {prof['year']} {prof['edu']})")
                    # v1.0.29：全欄位記錄（簡易模式以預設值補齊）
                    record = {"no": i, "code": code, "ts": ts}
                    for k, _, dv, _ in COMPLETE_FIELDS:
                        v = prof.get(k)
                        if v is None or v == "":
                            v = dv
                        record[k] = v
                    # 簡易 profile 的 testing 欄位 → 對應到 testing_habit
                    if "testing" in prof and (record.get("testing_habit") in (None, "", "否")):
                        record["testing_habit"] = prof["testing"]
                    self.results.append(record)
                    # v1.0.12 增量儲存 — 每筆完成立刻寫 Excel
                    self._incremental_save_excel()
                    self._flash_stats()  # v1.0.18 成功動畫
                else:
                    self.log(f"✗ 第 {i} 筆 失敗，跳過")
                self.pb.configure_value(i, len(pool))
                # v1.0.14 統計：依完成筆數計算平均 & ETA
                completed += 1
                elapsed = time.time() - run_start_ts
                avg = elapsed / completed if completed else 0
                total = len(pool)
                remaining = total - i
                eta_sec = int(avg * remaining)
                pct = (i / total * 100) if total else 0
                def _fmt(sec):
                    sec = max(0, int(sec))
                    h, r = divmod(sec, 3600); m, s = divmod(r, 60)
                    return f"{h}h{m:02d}m{s:02d}s" if h else f"{m}m{s:02d}s"
                # v1.0.16 預計完成時間（絕對時間 HH:MM）
                eta_at = (datetime.datetime.now() +
                          datetime.timedelta(seconds=eta_sec)).strftime("%H:%M:%S")
                # v1.0.19 完成%已在進度條內，這列只剩平均/已用/剩餘/預計完成
                self.stats_var.set(
                    f"平均 {avg:.1f} 秒/筆   |   已用 {_fmt(elapsed)}   |   "
                    f"剩餘 {_fmt(eta_sec)}   |   預計完成 {eta_at}")
                # v1.0.19 同步單筆模式結果顯示
                if len(pool) == 1 and code:
                    try: self.single_result_var.set(f"✓ 取到代碼：{code}")
                    except Exception: pass
                self.root.update_idletasks()
                if i < len(pool):
                    # 暫停期間不消耗 between 延遲（先等暫停解除）
                    while self.pause_evt.is_set() and not self.stop_evt.is_set():
                        time.sleep(0.5)
                    if not self.stop_evt.is_set():
                        dly.between()
        finally:
            worker.quit()
            self._save_excel()
            # v1.0.14 全部跑完（非中途停止）且勾選自動開啟 Excel
            done_naturally = not self.stop_evt.is_set()
            if done_naturally and self.auto_open_xlsx.get() and self._current_xlsx_path \
                    and os.path.exists(self._current_xlsx_path):
                self.log("✓ 全部完成，自動開啟 Excel…")
                try: os.startfile(self._current_xlsx_path)
                except Exception as e: self.log(f"⚠ 自動開啟失敗：{e}")
            # v1.0.12 停止鍵 → 存完自動關閉程式
            if self.stop_evt.is_set():
                self.log("✓ 已關閉瀏覽器、存檔完成 → 程式即將自動關閉")
                # v1.0.21 寫入 resume_state（給下次啟動續傳）
                try:
                    rs_state = {
                        "version": VERSION,
                        "saved_at": datetime.datetime.now().isoformat(),
                        "pool": self._current_pool,
                        "completed_index": self._completed_index,
                        "xlsx_path": self._current_xlsx_path,
                        "mode": self.mode_var.get(),
                    }
                    rs_path2 = os.path.join(OUTPUT_DIR, "resume_state.json")
                    with open(rs_path2, "w", encoding="utf-8") as f:
                        json.dump(rs_state, f, ensure_ascii=False, indent=2, default=str)
                    _hide_path(rs_path2)  # v1.0.43 L4
                    self.log(f"💾 已寫入續傳檔（剩 {len(self._current_pool) - self._completed_index} 筆）")
                except Exception as e:
                    self.log(f"⚠ 續傳檔寫入失敗：{e}")
                try: self._save_settings()
                except Exception: pass
                self.root.after(800, self._force_exit)
            else:
                self._finish()
                # v1.0.14 自然完成 + 勾選 → 同樣自動關閉程式（如使用者勾「完成後自動停止」）
                if done_naturally and self.auto_open_xlsx.get():
                    self.log("✓ 任務結束（已開 Excel），程式 3 秒後關閉…")
                    self.root.after(3000, self._force_exit)

    def _save_excel(self):
        """結束時最終確認儲存（增量儲存可能已經寫好了，這裡再確認一次）"""
        if not self.results:
            self.log("（無資料可存）")
            return
        try:
            self._incremental_save_excel()
            if self._current_xlsx_path and os.path.exists(self._current_xlsx_path):
                self.log(f"💾 已存 Excel → {self._current_xlsx_path}（共 {len(self.results)} 筆）")
            else:
                self.log("⚠ Excel 路徑異常，請手動檢查 number 資料夾")
        except Exception as e:
            self.log(f"⚠ 最終儲存失敗：{e}")

    def _force_exit(self):
        """v1.0.12：停止後完全退出進程（避免 Tk thread 殘留）"""
        try:
            self.root.destroy()
        except Exception:
            pass
        os._exit(0)

    def _finish(self):
        self.status(f"完成 — 成功 {len(self.results)} 筆")
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.pause_btn.config(state="disabled", text="⏸ 暫停")
        self._stop_spinner()
        self._set_settings_locked(False)  # v1.0.21 解鎖
        # v1.0.51 跑者抵達歡呼或正常結束
        try:
            if self.results and self.pb.value >= self.pb.maximum:
                self.pb.set_state('arrived')
            else:
                self.pb.set_state('idle')
        except Exception:
            pass
        try:
            if "尚未取號" in self.single_result_var.get() or "執行中" in self.single_result_var.get():
                if not self.results:
                    self.single_result_var.set("✗ 取號失敗，請看 log")
        except Exception:
            pass
        # v1.0.59 批次結束 → 上報雲端（精確產出代碼數）
        try: self._report_batch_done()
        except Exception: pass

    def _report_batch_done(self):
        """v1.0.59：批次結束精確上報該次 EXE 產出代碼數。
        每台機器累計就是 SUM(count) by hostname。背景 thread，失敗無聲。"""
        count = len(getattr(self, "results", []) or [])
        target = getattr(self, "_batch_target", 0)
        start_ts = getattr(self, "_batch_start_ts", 0)
        duration = int(time.time() - start_ts) if start_ts else 0
        # 判斷狀態：完成 vs 中斷
        if target > 0 and count >= target:
            status = "completed"
        elif self.stop_evt.is_set():
            status = "aborted"
        else:
            status = "partial"
        # count == 0 不送（沒意義）
        if count == 0 and status != "aborted":
            return
        info = _machine_info()

        def _bg():
            import urllib.request, urllib.error
            try:
                body = json.dumps({
                    "version": VERSION,
                    "count": count,
                    "duration_sec": duration,
                    "status": status,
                    **info,  # hostname/win_user/mac/os_ver
                }).encode("utf-8")
                req = urllib.request.Request(
                    CLOUD_AUTH_URL.rstrip("/") + "/report-batch",
                    data=body,
                    headers={
                        "Content-Type": "application/json;charset=utf-8",
                        "User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)",
                    },
                    method="POST",
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    resp.read()
            except Exception:
                pass  # 上報失敗不影響使用者
        threading.Thread(target=_bg, daemon=True).start()


# ── v1.0.39 雲端授權閘 ───────────────────────────────
# 啟動時連線 CLOUD_AUTH_URL/verify 驗證密碼。
# 完全不允許離線：連不到網路或 worker 回 ok=false 即拒絕。
# 管理員可在 cloud_auth/admin 改密碼或停用所有 EXE。
def _machine_info():
    """v1.0.56：取機器識別資訊，給 cloud_verify 與 _maybe_report_error 共用
    /verify 也帶上 hostname，admin UI 「7 天活躍機器」用 COUNT(DISTINCT hostname) 統計"""
    import socket, uuid
    info = {"hostname": "", "win_user": "", "mac": "", "os_ver": ""}
    try: info["hostname"] = (socket.gethostname() or "")[:64]
    except Exception: pass
    try: info["win_user"] = (os.environ.get("USERNAME") or os.environ.get("USER") or "")[:64]
    except Exception: pass
    try:
        mac_int = uuid.getnode()
        info["mac"] = ":".join(f"{(mac_int >> i) & 0xff:02x}" for i in (40, 32, 24, 16, 8, 0))
    except Exception: pass
    try:
        v = sys.getwindowsversion()
        info["os_ver"] = f"Windows {v.major}.{v.minor}.{v.build}"[:64]
    except Exception: pass
    return info


def cloud_verify(password):
    """回傳 (ok: bool, reason: str)。reason 直接顯示給使用者。
    v1.0.41：加 User-Agent，繞過 Cloudflare 對 Python-urllib 的 1010 阻擋
    v1.0.43：連線錯誤自動 retry 1 次（隔 1.5 秒），緩解現場 4G 抖動。
             伺服器明確拒絕（HTTP 4xx）不 retry，避免幫攻擊者放大試誤。
    v1.0.56：body 多帶 hostname（給 admin UI 算活躍機器數）"""
    import urllib.request
    import urllib.error
    info = _machine_info()
    body = json.dumps({
        "password": password,
        "hostname": info["hostname"],  # v1.0.56：給 audit COUNT(DISTINCT hostname) 用
    }).encode("utf-8")

    def _once():
        req = urllib.request.Request(
            CLOUD_AUTH_URL.rstrip("/") + "/verify",
            data=body,
            headers={
                "Content-Type": "application/json;charset=utf-8",
                "User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=CLOUD_AUTH_TIMEOUT) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            if data.get("ok"):
                return True, "", False  # ok, reason, transient
            return False, data.get("reason", "驗證失敗"), False
        except urllib.error.HTTPError as e:
            try:
                data = json.loads(e.read().decode("utf-8"))
                msg = data.get("reason", f"伺服器錯誤 HTTP {e.code}")
            except Exception:
                msg = f"伺服器錯誤 HTTP {e.code}"
            # 5xx 視為暫時性，4xx 不 retry
            transient = 500 <= e.code < 600
            return False, msg, transient
        except urllib.error.URLError as e:
            reason = getattr(e, "reason", str(e))
            return False, f"連線失敗：{reason}（請確認網路）", True
        except Exception as e:
            return False, f"請求失敗：{e}", True

    ok, reason, transient = _once()
    if ok:
        return True, ""
    if transient:
        time.sleep(1.5)
        ok, reason, _ = _once()
        if ok:
            return True, ""
    return False, reason


def show_password_gate(parent):
    """v1.0.39 每次啟動連線雲端驗證；無離線寬限。
    v1.0.40 修：parent 是 withdraw 過的，不能用 transient 否則視窗藏起來看不到。"""
    dlg = tk.Toplevel(parent)
    dlg.title("諮詢代碼取號小工具－登入")
    dlg.resizable(False, False)
    # 不用 transient，避免被 hidden parent 影響顯示
    dlg.grab_set()
    dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)

    result = {"ok": False}

    # v1.0.57：登入框隆重化 — bg 用主題色、字級放大、輸入框大、留白寬鬆
    BG = "#eef5fa"
    ACCENT = "#1565c0"
    ACCENT_DARK = "#0d47a1"

    # 頂部 accent 色 banner（讓視覺有重量）
    banner = tk.Frame(dlg, bg=ACCENT, height=6)
    banner.pack(fill="x")

    frm = tk.Frame(dlg, padx=44, pady=32, bg=BG)
    frm.pack(fill="both", expand=True)

    # 大標題（icon 跟文字並排）— v1.0.58 icon 改紅絲帶 🎗 (HIV 公益意象)
    head_fr = tk.Frame(frm, bg=BG)
    head_fr.pack(anchor="w")
    tk.Label(head_fr, text="🎗",
             font=("Segoe UI Emoji", 32),
             bg=BG, fg="#c62828").pack(side="left", padx=(0, 14))
    title_fr = tk.Frame(head_fr, bg=BG)
    title_fr.pack(side="left")
    tk.Label(title_fr, text="諮詢代碼取號小工具",
             font=("Microsoft JhengHei", 18, "bold"),
             bg=BG, fg=ACCENT).pack(anchor="w")
    tk.Label(title_fr, text="請輸入授權密碼以登入",
             font=("Microsoft JhengHei", 10),
             bg=BG, fg="#5a6577").pack(anchor="w", pady=(2, 0))

    # 分隔線
    tk.Frame(frm, bg="#cfd8dc", height=1).pack(fill="x", pady=(20, 14))

    # v1.0.58 版本資訊條（左：目前版本、右：雲端最新版本 / 檢查中…）
    ver_fr = tk.Frame(frm, bg=BG)
    ver_fr.pack(fill="x", pady=(0, 14))
    tk.Label(ver_fr, text=f"目前版本：v{VERSION}", bg=BG, fg="#37474f",
             font=("Consolas", 10)).pack(side="left")
    latest_var = tk.StringVar(value="雲端版本：檢查中…")
    latest_lbl = tk.Label(ver_fr, textvariable=latest_var, bg=BG, fg="#7d8696",
                           font=("Consolas", 10))
    latest_lbl.pack(side="right")

    # 密碼輸入區（label + Entry 大、留白多）
    tk.Label(frm, text="密碼", bg=BG, fg="#37474f",
             font=("Microsoft JhengHei", 11, "bold")).pack(anchor="w")

    # Entry 用一個內框包，模擬 border-radius 與較重的視覺
    entry_wrap = tk.Frame(frm, bg="#ffffff", highlightthickness=2,
                          highlightbackground="#cfd8dc",
                          highlightcolor=ACCENT)
    entry_wrap.pack(fill="x", pady=(8, 4))
    e1 = tk.Entry(entry_wrap, show="●", font=("Microsoft JhengHei", 16),
                  bg="#ffffff", fg="#1a2230",
                  relief="flat", bd=0,
                  insertbackground=ACCENT)
    e1.pack(fill="x", padx=14, ipady=10)

    msg = tk.Label(frm, text="", fg="#c62828", bg=BG,
                   font=("Microsoft JhengHei", 10), wraplength=420, justify="left")
    msg.pack(anchor="w", pady=(8, 0))

    busy = {"v": False}
    btn_login = None
    # v1.0.58 強制更新狀態
    update_state = {"required": False, "url": None, "filename": None, "latest": None}

    def do_login():
        if busy["v"]:
            return
        p = e1.get()
        if not p:
            msg.config(text="✗ 請輸入密碼")
            return
        busy["v"] = True
        msg.config(text="🌐 連線雲端驗證中…", fg="#1565c0")
        btn_login.config(state="disabled")
        e1.config(state="disabled")
        dlg.update_idletasks()
        try:
            ok, reason = cloud_verify(p)
        finally:
            busy["v"] = False
            try:
                e1.config(state="normal")
                btn_login.config(state="normal")
            except Exception:
                pass
        if ok:
            result["ok"] = True
            dlg.destroy()
        else:
            msg.config(text=f"✗ {reason}", fg="#c62828")
            e1.delete(0, tk.END)
            e1.focus()

    btnf = tk.Frame(frm, bg=BG)
    btnf.pack(fill="x", pady=(28, 0))

    # v1.0.58 強制更新處理：偵測新版時把「登入」改成「立即更新」
    def _do_force_update():
        if busy["v"]: return
        busy["v"] = True
        btn_login.config(state="disabled", text="下載中…")
        msg.config(text="🌐 正在下載新版 EXE，下載完請關閉本工具切換新版…", fg=ACCENT)
        dlg.update_idletasks()
        import urllib.request
        try:
            url = update_state["url"]
            filename = update_state["filename"] or f"HIV取號_v{update_state['latest']}.exe"
            exe_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
                      else os.path.dirname(os.path.abspath(__file__))
            new_path = os.path.join(exe_dir, filename)
            tmp = new_path + ".downloading"
            with urllib.request.urlopen(url, timeout=120) as resp:
                with open(tmp, "wb") as f:
                    while True:
                        chunk = resp.read(64 * 1024)
                        if not chunk: break
                        f.write(chunk)
            os.replace(tmp, new_path)
            msg.config(text=f"✅ 新版已下載：{filename}\n請關閉本工具，雙擊新 EXE 啟動",
                       fg="#2e7d32")
            btn_login.config(text="關閉", command=dlg.destroy, state="normal")
            try: os.startfile(exe_dir)  # 開資料夾讓使用者看到新檔
            except Exception: pass
        except Exception as e:
            msg.config(text=f"⚠ 下載失敗：{e}\n請改至 admin UI 手動下載", fg="#c62828")
            btn_login.config(state="normal")
        finally:
            busy["v"] = False

    # 取消（左）+ 登入（右），登入按鈕隆重大顆
    tk.Button(btnf, text="取消", width=10, command=dlg.destroy,
              font=("Microsoft JhengHei", 12),
              bg="#ffffff", fg="#37474f",
              activebackground="#eceff1", activeforeground="#37474f",
              relief="flat", bd=0, padx=18, pady=10, cursor="hand2").pack(side="left")
    btn_login = tk.Button(btnf, text="登 入", width=14, command=do_login,
              bg=ACCENT, fg="white",
              activebackground=ACCENT_DARK, activeforeground="white",
              font=("Microsoft JhengHei", 13, "bold"),
              relief="flat", bd=0, padx=22, pady=10, cursor="hand2")
    btn_login.pack(side="right")
    e1.focus()
    dlg.bind("<Return>", lambda e: do_login() if not update_state["required"] else _do_force_update())

    # v1.0.58/61 啟動時背景查 /version；watchdog 5 秒強制超時
    # done flag 防止 thread 與 watchdog 同時更新 UI 造成競態
    check_state = {"done": False}

    def _async_check_version():
        import urllib.request
        try:
            req = urllib.request.Request(
                CLOUD_AUTH_URL.rstrip("/") + "/version",
                headers={"User-Agent": f"HIV-Auth-Client/{VERSION} (Windows)"},
            )
            # v1.0.61：用 socket-level timeout 才確實會中斷（urllib 內部 readtimeout 有時不準）
            import socket
            socket.setdefaulttimeout(3)
            try:
                with urllib.request.urlopen(req, timeout=3) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
            finally:
                socket.setdefaulttimeout(None)
            if check_state["done"]: return  # 已被 watchdog 接管
            check_state["done"] = True
            if not data.get("ok"):
                dlg.after(0, lambda: latest_var.set("雲端版本：尚未上傳"))
                return
            latest = (data.get("latest") or "").strip()
            if not latest:
                return
            cmp_ = _semver_compare(latest, VERSION)

            def _apply_update_required():
                latest_var.set(f"雲端版本：v{latest}  ⚠ 必須更新")
                latest_lbl.config(fg="#c62828", font=("Consolas", 10, "bold"))
                msg.config(text=f"⚠ 偵測到新版 v{latest}（你目前 v{VERSION}），必須先更新才能繼續使用。",
                           fg="#c62828")
                e1.config(state="disabled")
                update_state["required"] = True
                update_state["latest"] = latest
                update_state["filename"] = data.get("filename")
                durl = data.get("download_url") or "/exe-download"
                if not durl.startswith("http"):
                    durl = CLOUD_AUTH_URL.rstrip("/") + durl
                update_state["url"] = durl
                btn_login.config(text="立即更新", command=_do_force_update,
                                  bg="#c62828", activebackground="#8b0000")

            def _apply_latest():
                latest_var.set(f"雲端版本：v{latest}  ✓ 已是最新")
                latest_lbl.config(fg="#2e7d32")

            if cmp_ > 0:
                dlg.after(0, _apply_update_required)
            else:
                dlg.after(0, _apply_latest)
        except Exception as e:
            if check_state["done"]: return
            check_state["done"] = True
            err_msg = str(e)[:40]
            dlg.after(0, lambda m=err_msg: latest_var.set(f"雲端版本：— （連線失敗）"))

    def _watchdog():
        if not check_state["done"]:
            check_state["done"] = True
            try: latest_var.set("雲端版本：— （超時 5 秒）")
            except Exception: pass

    threading.Thread(target=_async_check_version, daemon=True).start()
    dlg.after(5000, _watchdog)  # v1.0.61：5 秒 hard timeout 確保 UI 一定更新

    dlg.update_idletasks()
    w = dlg.winfo_reqwidth()
    h = dlg.winfo_reqheight()
    sw = dlg.winfo_screenwidth()
    sh = dlg.winfo_screenheight()
    dlg.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    # v1.0.40 強制把視窗推到最前面（避免被其他視窗蓋住或在背景看不見）
    dlg.deiconify()
    dlg.lift()
    dlg.attributes("-topmost", True)
    dlg.after(300, lambda: dlg.attributes("-topmost", False))
    dlg.focus_force()
    e1.focus_set()

    parent.wait_window(dlg)
    return result["ok"]


def _crash_log(exc_text):
    """v1.0.40：把 fatal error 寫到 EXE 旁邊的 crash_*.txt，方便除錯"""
    try:
        base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
               else os.path.dirname(os.path.abspath(__file__))
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        p = os.path.join(base, f"crash_{ts}.txt")
        with open(p, "w", encoding="utf-8") as f:
            f.write(f"VERSION {VERSION}\n")
            f.write(f"sys.executable: {sys.executable}\n")
            f.write(f"OUTPUT_DIR: {OUTPUT_DIR}\n")
            f.write(f"CLOUD_AUTH_URL: {CLOUD_AUTH_URL}\n")
            f.write("=" * 50 + "\n")
            f.write(exc_text)
        return p
    except Exception:
        return None


def main():
    try:
        log_path = init_logfile()
        root = tk.Tk()
        root.withdraw()  # 主視窗先藏，密碼通過才顯示
        try:
            root.iconbitmap(default="")
        except Exception:
            pass

        if not show_password_gate(root):
            root.destroy()
            sys.exit(0)

        root.deiconify()
        app = App(root)
        app._update_counts()
        app.log(f"📁 log 寫入：{log_path}")
        app.log(f"📁 輸出資料夾：{OUTPUT_DIR}")
        # v1.0.60：拿掉「舊版自動歸檔」「失敗快照」兩行 log（使用者沒在用、佔版面）

        # v1.0.55：背景檢查更新（thread-safe log 透過 root.after 切回主執行緒）
        def _safe_log(msg):
            try: root.after(0, lambda m=msg: app.log(m))
            except Exception: pass
        threading.Thread(
            target=check_and_download_update,
            args=(_safe_log,),
            daemon=True,
        ).start()

        root.mainloop()
    except SystemExit:
        raise
    except Exception:
        import traceback
        text = traceback.format_exc()
        crash_path = _crash_log(text)
        try:
            messagebox.showerror(
                "HIV 取號工具 — 啟動失敗",
                f"程式啟動時發生例外。\n\nlog 寫到：\n{crash_path or '（寫入失敗）'}\n\n"
                f"錯誤摘要：\n{text[-400:]}",
            )
        except Exception:
            pass


if __name__ == "__main__":
    main()
